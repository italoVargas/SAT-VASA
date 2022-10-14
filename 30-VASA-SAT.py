import calendar

from tkinter import *
import pywhatkit
import json
from collections import defaultdict
import xml.etree.ElementTree as ET
from PyPDF2 import PdfFileReader
import datetime
import os
import sys 
import pyautogui as PA
import time
import pandas as pd
import zipfile
import shutil
from difflib import SequenceMatcher as SM
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.chrome.options import Options

sys.path.append(os.path.abspath("/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON"))
from DatosClientes import *

class SAT():

    def __init__(self,
                NOMBRE="",
                path="",#ingreso o EGreso
                ANIOF=datetime.date.today().year,
                MESF=datetime.date.today().month,
                DIAF="",
                TipoConsulta=""):#Consulta 
        self.TEsperaMax = 10000
        self.TipoConsulta = TipoConsulta.upper()
        self.path=path.upper()
        
        self.NOMBRE = NOMBRE.upper()

        self.NombreCompleto=DIC_NOMBRES.get(self.NOMBRE)
        self.RFC=DIC_RFC.get(self.NOMBRE)
        self.EMAIL=DIC_MAIL.get(self.NOMBRE)
        self.PASS=DIC_PASS.get(self.NOMBRE)
        self.CP=DIC_CP.get(self.NOMBRE)
        #pass EFIRMA
        self.PASSEF=DIC_PASSEF.get(self.NOMBRE)
        #REgimenfiscal
        self.RF=DIC_RF.get(self.NOMBRE)
        #CLAVEPRODUCTO
        self.CLAVEPROD=DIC_CLAVEPROD.get(self.NOMBRE)        
          #direccionCOMPLETA
        self.DIRECCION=DIC_DIRECCION.get(self.NOMBRE)
        self.ANIOACT=datetime.date.today().year
        self.MESACT=datetime.date.today().month
        self.DIAACT=datetime.date.today().day
          #DIC CER y KEY 
        self.NOMCER=DIC_NOMCER.get(self.NOMBRE)
        self.NOMKEY=DIC_NOMKEY.get(self.NOMBRE)
        #DATOS DE FEHCA DE FACTURACION O CONSULTA 
        #no int
        self.ANIOF=ANIOF
        self.MESF=MESF
        self.DIAF=DIAF

        #1-IT-Fac-SAT
        self.PATHSYS1="/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT"

          #1-CLIENTES
        self.PATHSYS2="/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/1-CLIENTES"
        
        #pathEFIRMA
        self.ClientePathEFIRMA=str(f"{self.PATHSYS1}/EFIRMAS-CLIENTES/{self.NombreCompleto}")

        #pathRegistro ANUAL
        self.ClientePathRegAnual=str(f"{self.PATHSYS2}/1-ClientesPy/{self.RF}/{self.NombreCompleto}/{self.ANIOF}")

        #pathRegistro EXMPLE
        self.ClientePathRegEXMP=str(f"{self.PATHSYS1}/REGISTRO_RESICO-5.xlsx")

       # PATHSPARA ALMACENAR FACTURAS EGRESO CONSULTADAS
        self.ClientePathFE=str(f"{self.PATHSYS2}/1-ClientesPy/{self.RF}/{self.NombreCompleto}/{self.ANIOF}/{self.MESF}/FacturasEgreso/{str(datetime.datetime.now())}")

       # PATHSPARA ALMACENAR FACTURAS INGRESO CONSULTADAS
        self.ClientePathFI=str(f"{self.PATHSYS2}/1-ClientesPy/{self.RF}/{self.NombreCompleto}/{self.ANIOF}/{self.MESF}/FacturasIngreso/{str(datetime.datetime.now())}")

       # PATHSPARA ALMACENAR FACTURAS REALIZADAS A CLIENTES 
        self.ClientePathFR=str(f"{self.PATHSYS2}/1-ClientesPy/{self.RF}/{self.NombreCompleto}/{self.ANIOF}/{self.MESF}/FacturasRealizadas/{str(datetime.datetime.now())}")

        self.PathSolicitudFactura40=self.PATHSYS1+"/"+"1-SOLICITUD-FACTURAS-4.0"


        if self.path == "INGRESO":
            self.path = self.ClientePathFI
            if self.TipoConsulta ==  "F4":
                self.path = self.ClientePathFR

            try:
                shutil.rmtree(self.ClientePathFI)
                print("Old path eliminado.")
            except:
                pass

            self.TIPO= "INGRESO"
            print("Cambió  PATH a Ingreso; "+ self.path)





        elif self.path == "EGRESO":
            self.path = self.ClientePathFE
            if self.TipoConsulta ==  "F4":
                self.path = self.ClientePathFR


            try:
                #print("se intenta borrar")
                shutil.rmtree(self.ClientePathFE)
                print("Old path eliminado.")
            except:
                pass
            self.TIPO= "EGRESO"

            print("Cambió PATH a Egreso"+ self.path)


        elif self.TipoConsulta == "F3":
            self.path = self.ClientePathFR
            

        elif self.TipoConsulta ==  "F4":
            self.path = self.ClientePathFR

        elif self.TipoConsulta ==  "CONSULTA":
            pass


        else:
            print("---TIPO DE CONSULTA NO VÁLIDO---")
            quit()
            pass

            
            
        print("-----------------------------------------------------------------------------------------")
        print("---------------------------------------------------------")
        print("----------------------------------")
        print("__INIT__ Cliente: "+ self.NombreCompleto+"-"+str(self.TipoConsulta)+ "-"+str(self.TIPO)+ "-"+str(self.MESF)+"-"+str(self.ANIOF) )
        print("----------------------------------")
        print("---------------------------------------------------------")
        print("-----------------------------------------------------------------------------------------")

    def Web(self):#,path):
        self.TipoConsulta=str(self.TipoConsulta.upper())

        if self.TipoConsulta == "CONSULTA":# or "CONSULTA":
            self.URL="https://portalcfdi.facturaelectronica.sat.gob.mx/Consulta.aspx"

        elif self.TipoConsulta == "F4":# or "F4.0" or "FACTURA4.0":
            self.URL ="https://portal.facturaelectronica.sat.gob.mx/" 

        elif self.TipoConsulta == "F3":#or "FAC 3" or "FAC 3.3" or "FAC3.3" or "F3.3" or "FACTURA3.3":
            self.URL ="https://www.sat.gob.mx/aplicacion/26989/factura-electronica-en-mis-cuentas"
            #self.URL ="https://portalcfdi.facturaelectronica.sat.gob.mx/"
            #self.URL ="https://www.sat.gob.mx/aplicacion/82868/genera-tu-factura-electronica-de-forma-gratuita" 

        else:
            print("Operación Inválida.")
            pass
            quit()

        self.Opt = webdriver.ChromeOptions()
        self.Opt.add_argument("--incognito")
        self.profile = { 
        "download.default_directory": self.path, 
        "download.prompt_for_download": False
        }
        self.Opt.add_experimental_option("prefs", self.profile)
        #self.driver=webdriver.Safari()

        self.driver = webdriver.Chrome(f'{self.PATHSYS1}/chromedriver',chrome_options=self.Opt)

        #self.driver.set_window_size(500, 500)

        self.swait = WebDriverWait(self.driver, .5)
        self.lwait = WebDriverWait(self.driver, 1)
        #self.driver.maximize_window()
        #self.driver.set_window_size(500, 500)

        self.driver.get(self.URL)
        print("Cargó satisfactoriamente la web: "+self.URL)
        time.sleep(1)

    def quit(self):
        self.driver.quit()

    def log(self):  
        CuadRFC=self.driver.find_element(By.ID, "rfc")
        while CuadRFC!=None:

            self.driver.find_element(By.ID, "rfc").click()
            self.driver.find_element(By.ID, "rfc").send_keys(self.RFC)

            self.driver.find_element(By.ID, "password").click()
            self.driver.find_element(By.ID, "password").send_keys(self.PASS)

            Captcha = input("Captcha: ")
            self.driver.find_element(By.ID, "userCaptcha").click()
            self.driver.find_element(By.ID, "userCaptcha").send_keys(Captcha)
            
            self.driver.find_element(By.ID, "submit").click()

            try:
                time.sleep(1)
                CuadRFC=self.driver.find_element(By.ID, "rfc")
                time.sleep(.1)
                print("Introduce Nuevamente el CAPTCHA")

            except: 
                break
            
        print("Ingresamos Sistemas de Administración Tributaria")
    
    def logEF(self): 
        BtnEF = expected_conditions.presence_of_element_located((By.ID, "buttonFiel"))
        WebDriverWait(self.driver, self.TEsperaMax).until(BtnEF)
        time.sleep(.5)

        self.driver.find_element(By.ID, "buttonFiel").click()

        Btncer = expected_conditions.presence_of_element_located((By.ID, "btnCertificate"))
        WebDriverWait(self.driver, self.TEsperaMax).until(Btncer)
        time.sleep(.5)

        time.sleep(.5)
        self.driver.find_element(By.ID, "fileCertificate").send_keys(self.ClientePathEFIRMA+"/"+self.NOMCER)
        time.sleep(.5)
        self.driver.find_element(By.ID, "filePrivateKey").send_keys(self.ClientePathEFIRMA+"/" +self.NOMKEY)
        
        time.sleep(.5)
        self.driver.find_element(By.ID, "privateKeyPassword").click()
        self.driver.find_element(By.ID, "privateKeyPassword").send_keys(self.PASSEF)
        time.sleep(.5)

        self.driver.find_element(By.ID, "submit").click()
        time.sleep(1)

        try:
            err=self.driver.find_element(By.ID, "divError")
            print("error EFIRMA")

        except:
            print("Loging éxitoso con eFirma portable.-->"+str(self.NOMBRE))

    def Fac40(self,FGlobal="NO",Periodicidad="Mensual",Mes="",ano="",TipoFactura="Ingreso",DescDetallada="Venta",ProdServ="",UnidadMedida="Unidad de servicio", Cantidad="1",VUnitario="0",Descuento="0",ObjetoImpuesto="si",NumIdentif="00",MultiplicadorConcepto=1,Moneda="Peso Mexicano",RfcCliente="",NombreoRSCliente="",CpCliente="",RegFisCliente="",UsoFac="Gastos en general",FormaDePago="Efectivo", MetodoPago="Pago en una sola exhibición"):

        try:
            for fpy in os.listdir(self.PathSolicitudFactura40):
                try:
                    if fpy.endswith(".py"):
                        self.f=open(self.PathSolicitudFactura40+ "/"+str(fpy),"r")
                        self.textf=self.f.read()
                        self.TextF=self.textf
                        try:
                            self.DicTexF= json.loads(self.TextF)
                        except:
                            print("ERROR, Archivo .py en Solicitud de C.F.D.I ver:4.0 ")
                            break
                        print("Leyendo Solicitud de C.F.D.I: ", fpy )
                        #print(self.DicTexF,type(self.DicTexF))
        ## ASIGNACION VARIABLES DE ARHIVOS .py DE LA CARPETA DE SOLICITUD
                        for k,v in self.DicTexF.items():

                            if k== "FGlobal":
                                v=v.upper()
                                self.FGlobal = str(v)
                                print(k,v)

                            elif k== "Periodicidad":
                                v=v.upper()
                                self.Periodicidad = str(v)
                                print(k,v)

                            elif k== "Mes":
                                v=v.upper()
                                self.Mes = str(v)
                                print(k,v)

                            elif k== "Año":
                                self.ano = str(v)
                                print(k,v)

                            elif k== "TipoFactura":
                                v=v.upper()
                                self.TipoFactura = str(v)
                                print(k,v)

                            elif k== "DescDetallada":
                                self.DescDetallada = str(v)
                                print(k,v)

                            elif k== "ProdServ":
                                self.ProdServ = str(v)
                                print(k,v)

                            elif k== "UnidadMedida":
                                self.UnidadMedida = str(v)
                                print(k,v)

                            elif k== "Cantidad":
                                self.Cantidad = str(v)
                                print(k,v)

                            elif k== "VUnitario":
                                self.VUnitario = list(v)
                                print(k,self.VUnitario)

                                if self.RF != "Incorporación Fiscal":

                                    for i in self.VUnitario:
                                        VN=i/1.16
                                        indice=self.VUnitario.index(i)
                                        self.VUnitario[indice]=round(VN,2)
                                    print(k+"/1.16:",self.VUnitario)


                            elif k== "Descuento":
                                self.Descuento = str(v)
                                print(k,v)

                            elif k== "ObjetoImpuesto":
                                self.ObjetoImpuesto = str(v)
                                print(k,v)

                            elif k== "NumIdentif":
                                self.NumIdentif = str(v)
                                print(k,v)


                            elif k== "Moneda":
                                self.Moneda = str(v)
                                print(k,v)

                            elif k== "RfcCliente":
                                self.RfcCliente = str(v)
                                print(k,v)

                            elif k== "NommbreoRSCliente":
                                self.NombreoRSCliente = str(v)
                                print(k,v)

                            elif k== "CpCliente":
                                self.CpCliente = str(v)
                                print(k,v)

                            elif k== "RegFisCliente":
                                self.RegFisCliente = str(v)
                                print(k,v)

                            elif k== "UsoFac":
                                self.UsoFac = str(v)
                                print(k,v)

                            elif k== "FormaDePago":
                                self.FormaDePago = str(v)
                                print(k,v)

                            elif k== "MetodoPago":
                                self.MetodoPago = str(v)
                                print(k,v)

                            else:
                                pass

    #COMIENZO FACTURACION 4.0 SAT

                        element_present = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006Pregimen\']/i"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(element_present)
                        time.sleep(.1)

                        print(' <--------->')
                        print('Inicio de Facturación ver: 4.0 ')
                        time.sleep(7)



                        if self.FGlobal == "SI":
                            try:
                                time.sleep(1)
                                self.driver.find_element(By.ID, "135checkbox40")
                            except: 
                                setting = expected_conditions.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(setting)  

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/ul/li[2]/a").click()

                                avanzada = expected_conditions.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(avanzada)  

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span").click()
                                time.sleep(1)
                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span").click()
                                time.sleep(1)
                                self.driver.find_element(By.ID, "btnGuardar").click()
                                time.sleep(3)

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[2]/a").click()
                                time.sleep(.3)

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[2]/ul/li[1]/a").click()
                                time.sleep(3)

                                print("se habilito el boton de faactura global ")

                        else:
                            pass





                    
                        self.driver.find_element(By.XPATH, "//a[@id=\'E1350006Pregimen\']/i").click()
                        time.sleep(1)
                        self.driver.find_element(By.PARTIAL_LINK_TEXT, self.RF).click()
                        
                        #CP
                        time.sleep(.5)        

                        #tipoFactura INGRESO
                        if self.TipoFactura== "INGRESO":
                            print(self.TipoFactura)

                            try:
                                tx = expected_conditions.presence_of_element_located((By.XPATH, "/html/body/div[4]/form/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div[1]/div[2]/div[8]/div/div/span[3]"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(tx)
                                tx=self.driver.find_element(By.XPATH, "/html/body/div[4]/form/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div[1]/div[2]/div[8]/div/div/span[3]").text
                                if tx == "Ingreso":
                                    pass
                                    print("FINGRESOO PASSS")

                            except:
                                for n in range(0,1,1):

                                    tf = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006PtipodeFactura\']/i"))
                                    WebDriverWait(self.driver, self.TEsperaMax).until(tf)

                                    self.driver.find_element(By.XPATH, "//a[@id=\'E1350006PtipodeFactura\']/i").click()
                                    time.sleep(2)
                                
                                    self.driver.find_element(By.LINK_TEXT, "Ingreso").click()


                                    ep2 = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006Pregimen\']/i"))
                                    WebDriverWait(self.driver, self.TEsperaMax).until(ep2)

                                    print('Ya es C.F.D.I. tipo INGRESO')
                                    time.sleep(.1)

                            




                        elif self.TipoFactura== "EGRESO":

                            for n in range(0,1,1):
                                tf = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006PtipodeFactura\']/i"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(tf)

                                self.driver.find_element(By.XPATH, "//a[@id=\'E1350006PtipodeFactura\']/i").click()
                                time.sleep(2)
                            
                                self.driver.find_element(By.LINK_TEXT, "Egreso").click()


                                ep2 = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006Pregimen\']/i"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(ep2)

                                print('Ya es C.F.D.I tipo EGRESO')
                                time.sleep(.1)


                        else:
                            print("Seleccione el tipo del C.F.D.I  (INGRESO O EGRESO) -> POR DEFECTO INGRESO")
                            pass
                            


######################################
# ############
# ##                        ##############Fecha




                        print("passssoososos")
                        print(calendar.monthcalendar(int(self.ANIOF), int(self.MESF)))

                        CalendarioL=calendar.monthcalendar(int(self.ANIOF), int(self.MESF))
                        semana=0
                        print("passssoososos")
                        for i in CalendarioL:
                            semana+=1
                            try:
                                posicion=(i.index(int(self.DIAF)))+1
                                print("Indices de la fecha:")
                                print(semana,(i.index(int(self.DIAF)))+1)
                                break
                            except:
                                pass
 
                        fecha = expected_conditions.presence_of_element_located((By.XPATH, "//div[@id=\'A135row2\']/div[2]/div[6]/div[2]/a/i"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(fecha)
                        time.sleep(4)

                        self.driver.find_element(By.XPATH, "//div[@id=\'A135row2\']/div[2]/div[6]/div[2]/a/i").click()
                        time.sleep(1)



                        NumClick=int(self.MESACT)-int(self.MESF)
                        if NumClick >0:
                            for i in range(NumClick):
                                self.driver.find_element(By.CSS_SELECTOR,".datepicker-days .prev > .glyphicon").click()
                                time.sleep(.5)

                        elif NumClick <0:
                            NumClick=NumClick*(-1)
                            for i in range(NumClick):
                                self.driver.find_element(By.CSS_SELECTOR,".datepicker-days .next > .glyphicon").click()
                                time.sleep(.5)


                        else:
                            pass

                        try:
                            #######################    Seleccionar dia         Eje  Y               Eje X   
                            self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({semana}) > .day:nth-child({posicion})").click()

                        except:
                            self.driver.find_element(By.CSS_SELECTOR, ".today").click()
                            print("ERROR FECHA ")
                        time.sleep(1)



                        #FORMADE PAGO 

                        FP = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006PformadePago\']/i"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(FP) 
                    
                        print('Continúa Facturación ')
                    
                        time.sleep(3)
                        try:

                            self.driver.find_element(By.XPATH, "//a[@id=\'E1350006PformadePago\']/i").click()
                        except:
                            time.sleep(5)
                            self.driver.find_element(By.XPATH, "//a[@id=\'E1350006PformadePago\']/i").click()

                        FP = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006PformadePago\']/i"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(FP)   

                        time.sleep(2)

                        FPt = expected_conditions.presence_of_element_located((By.LINK_TEXT, self.FormaDePago))
                        WebDriverWait(self.driver, self.TEsperaMax).until(FPt)   
                        time.sleep(.1)

                        self.driver.find_element(By.LINK_TEXT, self.FormaDePago).click()


                        ##Método de pago

                        Mp = expected_conditions.presence_of_element_located((By.XPATH, "//a[@id=\'E1350006PmetododePago\']/i"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(Mp)  

                        time.sleep(1)

                        self.driver.find_element(By.XPATH, "//a[@id=\'E1350006PmetododePago\']/i").click()
                        MP = expected_conditions.presence_of_element_located((By.LINK_TEXT, self.MetodoPago))
                        WebDriverWait(self.driver, self.TEsperaMax).until(MP)  

                        time.sleep(1)
                        self.driver.find_element(By.LINK_TEXT, self.MetodoPago).click()
                        time.sleep(1)

                        print('Continúa Facturación ')


                        #MONEDA

                        Monedabtn = expected_conditions.presence_of_element_located((By.ID, "135textboxautocomplete32"))
                        WebDriverWait(self.driver, self.TEsperaMax).until(Monedabtn)  

                        moneda=self.driver.find_element(By.ID, "135textboxautocomplete32")
                        moneda.click()
                        time.sleep(.1)
                        moneda.clear()
                        time.sleep(.1)
                        moneda.send_keys(self.Moneda)
                        time.sleep(1.5)
                        moneda.send_keys(Keys.DOWN)
                        time.sleep(1)
                        moneda.send_keys(Keys.TAB)
                        time.sleep(1)


                        # PERIODICIDAD

                        if self.FGlobal == "SI":
                            try:
                                time.sleep(1)
                                self.driver.find_element(By.ID, "135checkbox40").click()
                            except: 
                                setting = expected_conditions.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(setting)  

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/ul/li[2]/a").click()

                                avanzada = expected_conditions.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(avanzada)  

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span").click()
                                time.sleep(1)
                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[3]/a/span").click()
                                time.sleep(1)
                                self.driver.find_element(By.ID, "btnGuardar").click()
                                time.sleep(3)

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[2]/a").click()
                                time.sleep(.3)

                                self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/ul/li[2]/ul/li[1]/a").click()
                                time.sleep(3)

                                print("notermino")

                            time.sleep(1)

                            self.driver.find_element(By.ID, "135select45").click()
                            time.sleep(1)

                            FG = expected_conditions.presence_of_element_located((By.XPATH, "//option[. = 'Diario']"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(FG)  

                            time.sleep(.5)
                            dropdown = self.driver.find_element(By.ID, "135select45")
                            if self.Periodicidad == "DIARIO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Diario']").click()
                            elif self.Periodicidad == "SEMANAL":
                                dropdown.find_element(By.XPATH, "//option[. = 'Semanal']").click()
                            elif self.Periodicidad == "QUINCENAL":
                                dropdown.find_element(By.XPATH, "//option[. = 'Quincenal']").click()
                            elif self.Periodicidad == "MENSUAL":
                                dropdown.find_element(By.XPATH, "//option[. = 'Mensual']").click()

                            else:
                                print("ERROR  Período")
                                pass

                            self.driver.find_element(By.ID, "135select46").click()
                            mes = expected_conditions.presence_of_element_located((By.XPATH, "//option[. = 'Enero']"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(mes)  

                            time.sleep(.5)
                            dropdown = self.driver.find_element(By.ID, "135select46")
                            if self.Mes == "ENERO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Enero']").click()
                            elif self.Mes == "FEBRERO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Febrero']").click()
                            elif self.Mes == "MARZO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Marzo']").click()
                            elif self.Mes == "ABRIL":
                                dropdown.find_element(By.XPATH, "//option[. = 'Abril']").click()
                            elif self.Mes == "MAYO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Mayo']").click()
                            elif self.Mes == "JUNIO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Junio']").click()
                            elif self.Mes == "JULIO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Julio']").click()
                            elif self.Mes == "AGOSTO":
                                dropdown.find_element(By.XPATH, "//option[. = 'Agosto']").click()
                            elif self.Mes == "SEPTIEMBRE":
                                dropdown.find_element(By.XPATH, "//option[. = 'Septiembre']").click()
                            elif self.Mes == "OCTUBRE":
                                dropdown.find_element(By.XPATH, "//option[. = 'Octubre']").click()
                            elif self.Mes == "NOVIEMBRE":
                                dropdown.find_element(By.XPATH, "//option[. = 'Noviembre']").click()
                            elif self.Mes == "DICIEMBRE":
                                dropdown.find_element(By.XPATH, "//option[. = 'Diciembre']").click()

                            else:
                                print("ERROR Mes")
                                pass
                            #año 
                            time.sleep(.5)

                            self.driver.find_element(By.ID, "135textbox47").click()
                            time.sleep(.5)

                            self.driver.find_element(By.ID, "135textbox47").clear()
                            time.sleep(.5)

                            self.driver.find_element(By.ID, "135textbox47").send_keys(self.ano)

                        else:
                            print("No es C.F.D.I. GOBAL")
                            pass


                        #CLICK RFC PUB GRAL .........

                        if self.FGlobal=="SI":
                            
                            otro = expected_conditions.presence_of_element_located((By.ID, "135textboxautocomplete55"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(otro)  

                            self.driver.find_element(By.ID, "135textboxautocomplete55").click()
                            self.driver.find_element(By.ID, "135textboxautocomplete55").clear()
                            self.driver.find_element(By.ID, "135textboxautocomplete55").send_keys("XAXX010101000")
                            time.sleep(1)
                            self.driver.find_element(By.ID, "135textboxautocomplete55").send_keys(Keys.ARROW_DOWN)
                            time.sleep(1)
                            self.driver.find_element(By.ID, "135textboxautocomplete55").send_keys(Keys.ENTER)
                            time.sleep(1)




                            self.driver.find_element(By.ID, "135textboxautocomplete66").click()
                            time.sleep(.7)

                            self.driver.find_element(By.ID, "135textboxautocomplete66").clear()
                            time.sleep(.1)

                            self.driver.find_element(By.ID, "135textboxautocomplete66").send_keys("Sin efectos fiscales")
                            time.sleep(.7)
                            self.driver.find_element(By.ID, "135textboxautocomplete66").send_keys(Keys.ARROW_DOWN)
                            time.sleep(.1)
                            self.driver.find_element(By.ID, "135textboxautocomplete66").send_keys(Keys.ENTER)
                            time.sleep(1)
                            self.driver.find_element(By.ID, "135textboxautocomplete66").send_keys(Keys.PAGE_DOWN)
                            time.sleep(2)


                        elif self.FGlobal=="NO":

                            #RFC CLIENTE frecuente "OTRO"
                            otro = expected_conditions.presence_of_element_located((By.ID, "135textboxautocomplete55"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(otro)  

                            CF= self.driver.find_element(By.ID, "135textboxautocomplete55")#.click()
                            time.sleep(.5)
                            CF.click()
                            time.sleep(.5)
                            CF.clear()
                            time.sleep(.5)

                            CF.send_keys("Otro")
                            time.sleep(1)
                            CF.send_keys(Keys.ARROW_DOWN)

                            time.sleep(.5)
                            CF.send_keys(Keys.ENTER)
                            time.sleep(.5)
                            rfcc = expected_conditions.presence_of_element_located((By.ID, "135textbox59"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(rfcc)


                            #RFC CLIENTE
                            try:
                                self.driver.find_element(By.ID, "135textbox59").click()

                            except:
                                rfcc = expected_conditions.presence_of_element_located((By.ID, "135textbox59"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(rfcc)  
                                time.sleep(3)
                                self.driver.find_element(By.ID, "135textbox59").click()

                            time.sleep(.5)

                            self.driver.find_element(By.ID, "135textbox59").send_keys(self.RfcCliente)
                            time.sleep(1)


                                #Nombre o Razon social 
                            NRsc = expected_conditions.presence_of_element_located((By.ID, "135textbox60"))
                            WebDriverWait(self.driver, self.TEsperaMax).until(NRsc)  
                            time.sleep(.1)    
                            try:
                                NRS=self.driver.find_element(By.ID, "135textbox60")
                                NRS.click()
                            except:
                                time.sleep(5)
                                NRS=self.driver.find_element(By.ID, "135textbox60")
                                NRS.click()

                            time.sleep(.5)
                            NRS.clear()
                        
                            time.sleep(.5)

                            self.driver.find_element(By.ID, "135textbox60").send_keys(self.NombreoRSCliente)

                            time.sleep(1)


                                #cpCLiente
                            self.driver.find_element(By.ID, "135textbox61").click()
                            time.sleep(.5)
                            self.driver.find_element(By.ID, "135textbox61").clear()
                            time.sleep(.5)
                            self.driver.find_element(By.ID, "135textbox61").send_keys(self.CpCliente)
                            time.sleep(1)

                            ## REGIMEN FISCAL DEL CLIENTE

                            BRefFC=self.driver.find_element(By.ID, "135textboxautocomplete62")
                            BRefFC.click()
                            time.sleep(.5)
                            BRefFC.clear()
                            time.sleep(.5)
                            BRefFC.send_keys(self.RegFisCliente)
                            time.sleep(2)
                            BRefFC.send_keys(Keys.ARROW_DOWN)
                            time.sleep(.5)
                            BRefFC.send_keys(Keys.ARROW_DOWN)
                            time.sleep(.5)
                            BRefFC.send_keys(Keys.ENTER)

                            time.sleep(.5)

                                #uso de la fac
                            
                            BUsFac0=self.driver.find_element(By.ID, "135textboxautocomplete71")
                            BUsFac0.click()
                            BUsFac0.clear()
                            time.sleep(.5)

                            BUsFac0.send_keys(self.UsoFac)
                            time.sleep(2)

                            BUsFac0.send_keys(Keys.DOWN)
                            time.sleep(1)

                            BUsFac0.send_keys(Keys.ENTER)


                            time.sleep(1)

                        else:
                            print("Factura global Val=ERR")
                            pass

                # CONCEPTOS
                        con=0
                        for vunit in self.VUnitario:# range(len(self.VUnitario)):
                            try:
                                con+=1
                                try:
                                    time.sleep(1)

                                    self.driver.find_element(By.XPATH, "(//button[@type=\'button\'])[17]").click()
                                    print("añadiendo concepto: "+str(con)+" de: "+str(len(self.VUnitario)))

                                except:
                                    agreg = expected_conditions.presence_of_element_located((By.XPATH,  "(//button[@type=\'button\'])[17]"))
                                    WebDriverWait(self.driver, self.TEsperaMax).until(agreg)  

                                    self.driver.find_element(By.XPATH, "(//button[@type=\'button\'])[17]").click()
                            
                                time.sleep(1)
                        
                                #descripcion detallada
                                self.driver.find_element(By.ID, "135textboxautocomplete112").click()
                                time.sleep(.5)      
                                self.driver.find_element(By.ID, "135textboxautocomplete112").clear()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textboxautocomplete112").send_keys(self.DescDetallada)
                                time.sleep(1)

                                #prod o servicio
                                self.driver.find_element(By.ID, "135textboxautocomplete118").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textboxautocomplete118").send_keys(self.ProdServ)
                                time.sleep(3)
                                self.driver.find_element(By.ID, "135textboxautocomplete118").send_keys(Keys.DOWN)
                                time.sleep(.5)
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textboxautocomplete118").send_keys(Keys.TAB)
                                time.sleep(.5)


                                #unidad de medida
                                self.driver.find_element(By.ID, "135textboxautocomplete122").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textboxautocomplete122").clear()
                                time.sleep(.5)

                                self.driver.find_element(By.ID, "135textboxautocomplete122").send_keys(self.UnidadMedida)
                                time.sleep(2)
                                self.driver.find_element(By.ID, "135textboxautocomplete122").send_keys(Keys.DOWN)
                                time.sleep(.6)
                                self.driver.find_element(By.ID, "135textboxautocomplete122").send_keys(Keys.ENTER)
                                time.sleep(.5)


                                #cantidad
                                self.driver.find_element(By.ID, "135textbox113").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox113").clear()
                                time.sleep(.5)

                                self.driver.find_element(By.ID, "135textbox113").send_keys(self.Cantidad)
                                time.sleep(1)

                                #Valor Unitario 
                                self.driver.find_element(By.ID, "135textbox119").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox119").clear()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox119").send_keys(str(vunit))
                                time.sleep(1)

                                #Descuento
                                self.driver.find_element(By.ID, "135textbox128").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox128").send_keys(self.Descuento)
                                time.sleep(.5)
                        


                                #obj de impuestos

                                BObjImp= self.driver.find_element(By.ID, "135select114")#.click()
                                BObjImp.click()

                                imp = expected_conditions.presence_of_element_located((By.XPATH, "//option[. = 'Sí objeto de impuesto.']"))
                                WebDriverWait(self.driver, self.TEsperaMax).until(imp)  

                                if self.ObjetoImpuesto== "no":
                                    dropdown = self.driver.find_element(By.ID, "135select114")
                                    dropdown.find_element(By.XPATH, "//option[. = 'No objeto de impuesto.']").click()
                                elif self.ObjetoImpuesto== "si":
                                    print("SIENTROA SIOBJETODEIMPUESOTOO")
                                    
                                    BObjImp.click()
                                    time.sleep(1)
                                    BObjImp.send_keys(self.ObjetoImpuesto)
                                    time.sleep(1)
                                    BObjImp.send_keys(Keys.ENTER)
                                    time.sleep(1)



                                elif self.ObjetoImpuesto== "Sí objeto de impuesto y no obligado al desglose.":
                                    dropdown = self.driver.find_element(By.ID, "135select114")
                                    dropdown.find_element(By.XPATH, "//option[. = 'Sí objeto de impuesto y no obligado al desglose.']").click()

                                else:
                                    pass


                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox120").click()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox120").clear()
                                time.sleep(.5)
                                self.driver.find_element(By.ID, "135textbox120").send_keys(self.NumIdentif)
                                time.sleep(.5)

                                if self.RF == "Régimen Simplificado de Confianza" or self.RF == "Personas Físicas con Actividades Empresariales y Profesionales" and self.RegFisCliente == "General de Ley Personas Morales":

                                    #Impuestos RETENIDOS

                                    impSug = expected_conditions.presence_of_element_located((By.ID, "135checkbox137"))
                                    WebDriverWait(self.driver, self.TEsperaMax).until(impSug)  


                                    self.driver.find_element(By.ID, "135checkbox137").click()
                                    time.sleep(.5)

                                    impRetISR = expected_conditions.presence_of_element_located((By.ID, "135checkbox168"))
                                    WebDriverWait(self.driver, self.TEsperaMax).until(impRetISR)  

                                    self.driver.find_element(By.ID, "135checkbox168").click()
                                    time.sleep(1.5)


                                    self.driver.find_element(By.ID, "135textboxautocomplete171").click()
                                    time.sleep(.5)

                                    self.driver.find_element(By.ID, "135textboxautocomplete171").clear()
                                    time.sleep(.5)
                                    self.driver.find_element(By.ID, "135textboxautocomplete171").send_keys("1.25%")
                                    time.sleep(.5)

                                    #self.driver.find_element(By.ID, "135textboxautocomplete171").send_keys(Keys.ENTER)
                                    self.driver.find_element(By.ID, "135textboxautocomplete171").send_keys(Keys.TAB)
                                    time.sleep(.5)
                                    self.driver.find_element(By.ID, "135textboxautocomplete171").send_keys(Keys.TAB)


                                time.sleep(1)


                                self.driver.find_element(By.ID, "guardarEditar1350001").click()
                                time.sleep(3)
                                try:
                                    self.driver.find_element(By.LINK_TEXT, "Aceptar").click()
                                except:
                                    pass


                            except:
                                print("###############        ALGO SALIO MAL            ###########, Intentando con otro concepto")
                                time.sleep(1)
                                pass



                        #guardar toda la factura y sellar 
                        self.driver.find_element(By.LINK_TEXT, "Guardar").click()
                        time.sleep(3)
                        try:
                            self.driver.find_element(By.LINK_TEXT, "Aceptar").click()
                        except:
                            pass

                        time.sleep(5)
                        PA.scroll(-1)

                        try:
                            print("Término de facturación, Validación y Sellado en 10 segundos ")
                            time.sleep(15)

                            self.driver.find_element(By.LINK_TEXT, "Sellar").click()
                            time.sleep(2)
                            self.driver.find_element(By.ID, "privateKeyPassword").click()
                            time.sleep(0.5)
                            self.driver.find_element(By.ID, "privateKeyPassword").send_keys(self.PASSEF)

                            time.sleep(0.5)
                            self.driver.find_element(By.ID, "privateKey").send_keys(self.ClientePathEFIRMA+"/" +self.NOMKEY)

                            time.sleep(1)
                            self.driver.find_element(By.ID, "certificate").send_keys(self.ClientePathEFIRMA+"/"+self.NOMCER)

                            time.sleep(0.5)
                            self.driver.find_element(By.ID, "btnValidaOSCP").click()

                            time.sleep(1)
                            self.driver.find_element(By.ID, "btnFirmar").click()
                            time.sleep(3)


                            pass


                        except:
                            print("nosepudo sellar")
                            time.sleep(2)
                            #self.driver.find_element(By.XPATH, "(//button[@type=\'button\'])[25]")
                            self.driver.find_element(By.XPATH, "(//button[@type=\'button\'])[25]").click()
                            print("nose sello hay error revisar ")
                            input("NOSE SELLO")
            #shutil.copyfile(self.ClientePathRegEXMP, self.ClientePathRegAnual+"/"+self.NombreCompleto+str(self.ANIOACT)+".xlsx")

                        shutil.copy(self.PathSolicitudFactura40+"/"+str(fpy),self.PathSolicitudFactura40+"/F-Realizadas/"+str(datetime.datetime.now)+"/"+str(fpy))
                        time.sleep(.1)
                        os.remove(self.PathSolicitudFactura40+"/"+str(fpy))
                        print("Termino FAC 4.0  SE REMOVIO ARCHIVO: "+ str(fpy))

######################################
################
################
                    else:
                        print("Solicitud de C.F.D.I. 4.0 debe estar con extensión  '.py' - Verificar")
                        pass

                except:
                    print("Ocurrió Error " + str(fpy))
                    input("INPUT:      Ocurrió Error " + str(fpy))
                    
                    pass

        except:
            print("NO SE PUDO FACTURA 4.0")
            time.sleep(5)
            pass

    def fac3I(self):
        element_present = expected_conditions.presence_of_element_located((By.CLASS_NAME, "actionButton"))
        WebDriverWait(self.driver, self.TEsperaMax).until(element_present)
        time.sleep(1)

        self.driver.find_element(By.CLASS_NAME, "actionButton").click()
        time.sleep(2)

        FFacil=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/4.png', confidence=0.9)
        PA.moveTo(FFacil)

        time.sleep(1)
        FFacil33=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/4-1.png', confidence=0.9)
        PA.click(FFacil33)
		
        time.sleep(3)
	
        rfc=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/5.png', confidence=0.9)
        PA.click(rfc)
        PA.write(self.RFC)

        time.sleep(1)
        pas=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/5.1.png', confidence=0.9)
        PA.click(pas)
        PA.write(self.PASS)
    
        time.sleep(1)

        captcha=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/5.2.png', confidence=0.9)
        PA.click(captcha)
        C=input('Ingresa el CAPTCHA: ')
        C=C.upper()
        PA.click(captcha)
        PA.write(C)
			
        PA.scroll(-5)
		
        enviar=PA.locateCenterOnScreen('/Users/italovargas/Library/CloudStorage/GoogleDrive-vargasserrano.asociados@gmail.com/My Drive/1-Contabilidad/2-SGC/PYTHON/1-IT-Fac-SAT/P/5.3.png', confidence=0.9)
        PA.click(enviar)

    def Consultar(self):
        self.DIAF=str(self.DIAF)
        self.MESF = str(self.MESF)
        self.ANOF = str(self.ANIOF)
        self.txtano= f"//option[. = '{self.ANIOF}'] "
        self.txtmes= f"//option[. = '{self.MESF}'] "
        self.txtdia= f"//option[. = '{self.DIAF}']"

        #Consultar Facturas Emitidas
        if self.TIPO == "INGRESO":

            CFacE = expected_conditions.presence_of_element_located((By.LINK_TEXT, "Consultar Facturas Emitidas"))
            WebDriverWait(self.driver, self.TEsperaMax).until(CFacE)
            time.sleep(1)

            self.driver.find_element(By.LINK_TEXT, "Consultar Facturas Emitidas").click()

            time.sleep(1)
             #Por FECHA

            BFecha = expected_conditions.presence_of_element_located((By.ID, "ctl00_MainContent_RdoFechas"))
            WebDriverWait(self.driver, self.TEsperaMax).until(BFecha)
            time.sleep(.1)


            self.driver.find_element(By.ID, "ctl00_MainContent_RdoFechas").click()
            time.sleep(5)

            BFecha = expected_conditions.presence_of_element_located((By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2"))
            WebDriverWait(self.driver, self.TEsperaMax).until(BFecha)
            time.sleep(.1)

            # Valores Anuales Coordenadas día primero y día último

            if self.MESF == "01":

                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)


                time.sleep(0.1)

                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[7]').click()
                time.sleep(0.5)

            #fecha FINAL

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.5)
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[8]/td[2]').click()
                time.sleep(0.5)


            elif self.MESF == "02":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)


                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[3]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[2]').click()
                time.sleep(0.5)




            elif self.MESF == "03":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)


                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[3]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[5]').click()
                time.sleep(0.5)




            elif self.MESF == "04":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)
                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[6]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[7]').click()
                time.sleep(0.5)




            elif self.MESF == "05":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[8]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[8]/td[3]').click()
                time.sleep(0.5)


            elif self.MESF == "06":

                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[4]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[5]').click()
                time.sleep(0.5)





            elif self.MESF == "07":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[6]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[8]/td[1]').click()
                time.sleep(0.5)


            elif self.MESF == "08":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[2]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[4]').click()
                time.sleep(0.5)


            elif self.MESF == "09":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
                 
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)
#INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[5]').click()
                time.sleep(.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(.5)
#fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[6]').click()
                time.sleep(.5)


            elif self.MESF == "10":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

#INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[7]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
#fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[8]/td[2]').click()
                time.sleep(0.5)





            elif self.MESF == "11":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

#INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[3]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[4]').click()
                time.sleep(0.5)


            elif self.MESF == "12":
                veces=int(self.MESACT)-int(self.MESF) 
                print(veces)
                #es una matriz de 3 a 8"Y" y de 1 a 7 "x"
  
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaInicial2_BtnFecha2").click()
                time.sleep(0.1)
                for i in range(veces):
                    self.driver.find_element(By.XPATH, "/html/body/div/table/tbody/tr[1]/td[1]/button").click()
                    time.sleep(0.5)

                                                                                         #INICIO
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[3]/td[5]').click()
                time.sleep(0.5)

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFechaFinal2_BtnFecha2").click()
                time.sleep(0.1)
                                                                                    #fecha FINAL
                self.driver.find_element(By.XPATH, '//*[@id="datepicker"]/table/tbody/tr[7]/td[7]').click()
                time.sleep(0.5)

            else: 
                pass

            self.driver.find_element(By.ID, "ctl00_MainContent_BtnBusqueda").click()
            time.sleep(1)
        
            try:
                meta=expected_conditions.presence_of_element_located((By.ID, "ListaFolios"))
                WebDriverWait(self.driver,15).until(meta)
                pass
            except:          
                meta=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, ".subtitle"))
                WebDriverWait(self.driver,15).until(meta)

                meta=self.driver.find_element(By.CSS_SELECTOR, ".subtitle").text.strip()
                if meta=="No existen registros que cumplan con los criterios de búsqueda ingresados, intentar nuevamente.":
                    print(meta)
                    print("NO HAY C.F.D.I.")
                    self.driver.quit()

                else:
                    pass

            meta=expected_conditions.presence_of_element_located((By.ID, "ListaFolios"))
            WebDriverWait(self.driver,15).until(meta)

            time.sleep(.5)
            for i in range(2,1000):

                try:

                    me=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga"))
                    WebDriverWait(self.driver,2).until(me)
                        
                    time.sleep(.5)
                    VentanaPrincipal=self.driver.window_handles[0]
                    print(VentanaPrincipal)
                    time.sleep(1)
#id=BtnVerDetalle     id=BtnDescarga          id=BtnRI
                    self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga").click()
                    time.sleep(2)

                    VentanaDescarga=self.driver.window_handles[1]
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaDescarga)
                    print(VentanaDescarga)

                    time.sleep(.5)

                    self.driver.close()
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaPrincipal)
                    time.sleep(.5)
                    print("CFDI XML Descargada  NUEMERO: " ,i-1)
                    time.sleep(1)

#Descarga del   PDF
                    self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnRI").click()
                    time.sleep(2)

                    VentanaDescarga=self.driver.window_handles[1]
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaDescarga)
                    print(VentanaDescarga)

                    time.sleep(.5)

                    self.driver.close()
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaPrincipal)
                    time.sleep(.5)
                    print("CFDI PDF Descargada  NUEMERO: " ,i-1)
                    time.sleep(.5)






                    #self.driver.find_element(By.LINK_TEXT,"»").click()
                    

                      
                except:
                    try:
                        txtfac=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) > td:nth-child(2) > span"))
                        WebDriverWait(self.driver,.5).until(txtfac)

                        if txtfac:
                            print("existe CFDI:", i-1, "pero no Esta el boton de descarga : CANCELADA ")

                    except:

                        print("No existe  CFDI:", i-1)
                        pass

                    

                if (i-1)%15==0:
                    try: 
                        self.driver.find_element(By.LINK_TEXT, "»").click()
                        time.sleep(.5)
                        self.driver.find_element(By.XPATH, '//body' ).send_keys(Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP)
                        #self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) > td:nth-child(2) > span").send_keys(Keys.ARROW_UP)
                        me=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga"))
                        WebDriverWait(self.driver,1).until(me)
            
                        print("Siguiente  página") 

                        time.sleep(.5)

                    except:
                        print("No hay más CFDI")
                        break
                              
            print("-----> Terminó Descarga de  CFDI de manera individual")


            time.sleep(1)

            self.driver.find_element(By.ID, "ctl00_MainContent_BtnMetadata").click()
            time.sleep(1)

            #nombre del archivo zip a descargar  y comparar
            meta=expected_conditions.presence_of_element_located((By.XPATH , '//*[@id="dvAlert"]/div/strong'))
            WebDriverWait(self.driver, self.TEsperaMax).until(meta)
            
            tmeta=self.driver.find_element(By.CSS_SELECTOR, ".alert-success").text.strip()
            tmeta=tmeta[137:173]

            print(tmeta)

            time.sleep(.5)

            self.driver.get("https://portalcfdi.facturaelectronica.sat.gob.mx/Consulta.aspx")
        #self.driver.find_element(By.LINK_TEXT, "Inicio").click()
            time.sleep(1)

            rd=expected_conditions.presence_of_element_located((By.LINK_TEXT, "Recuperar Descargas de CFDI"))
            WebDriverWait(self.driver, self.TEsperaMax).until(rd)
            time.sleep(.5)

            self.driver.find_element(By.LINK_TEXT, "Recuperar Descargas de CFDI").click()
            time.sleep(.5)
            
            time.sleep(.5)

            try:
                zipc=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)"))
                WebDriverWait(self.driver, 10).until(zipc)
                time.sleep(.5)
            except:
                self.driver.refresh()
            zi=0
            try:
                zi=self.driver.find_element(By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)").text.strip()

            except:
                self.driver.refresh()

            contador=0

            while zi != tmeta:

                try:
                    self.driver.refresh()
                    zipc=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)"))
                    WebDriverWait(self.driver, 5).until(zipc)
                    zi=self.driver.find_element(By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)").text.strip()

                    time.sleep(.5)
                except:
                    self.driver.refresh()
                self.driver.refresh()
                time.sleep(5)
                contador=contador+1
                print(tmeta+"-- Mes:"+self.MESF+"Contador: "+str(contador))
                print(zi)

            if zi == tmeta:
                print("aparece el mismo que descargaste se procedea descargar ")
                pass
            else:
                pass
            
            descargar=expected_conditions.presence_of_element_located((By.ID, "BtnDescarga"))
            WebDriverWait(self.driver, self.TEsperaMax).until(descargar)
            time.sleep(.5)

            self.driver.find_element(By.ID, "BtnDescarga").click()
            time.sleep(3)
        
        elif self.TIPO == "EGRESO":

            rec = expected_conditions.presence_of_element_located((By.LINK_TEXT, "Consultar Facturas Recibidas"))
            WebDriverWait(self.driver, self.TEsperaMax).until(rec)
            time.sleep(.5)

            self.driver.find_element(By.LINK_TEXT, "Consultar Facturas Recibidas").click()
            time.sleep(.5)



            #Por FECHA
            self.driver.find_element(By.ID, "ctl00_MainContent_RdoFechas").click()
            time.sleep(2)

            #Selección de Fecha
            ano = expected_conditions.presence_of_element_located((By.ID, "DdlAnio"))
            WebDriverWait(self.driver, self.TEsperaMax).until(ano)
            time.sleep(.5)
            
            self.driver.find_element(By.ID, "DdlAnio").click()

            esp = expected_conditions.presence_of_element_located((By.ID, "DdlAnio"))
            WebDriverWait(self.driver, self.TEsperaMax).until(esp)
            time.sleep(.5)

            Sano = expected_conditions.presence_of_element_located((By.XPATH, self.txtano ))
            WebDriverWait(self.driver, self.TEsperaMax).until(Sano)

            time.sleep(.5)

            self.driver.find_element(By.XPATH,self.txtano).click()
        
            time.sleep(.5)

        #mess

            SMES = expected_conditions.presence_of_element_located((By.ID, "ctl00_MainContent_CldFecha_DdlMes"))
            WebDriverWait(self.driver, self.TEsperaMax).until(SMES)

            self.driver.find_element(By.ID, "ctl00_MainContent_CldFecha_DdlMes").click()
            time.sleep(.5)

            SMES = expected_conditions.presence_of_element_located((By.XPATH, self.txtmes))
            WebDriverWait(self.driver, self.TEsperaMax).until(SMES)

            self.driver.find_element(By.XPATH, self.txtmes).click()
        
            time.sleep(.5)

            #Día
            if 0 < int(self.DIAF) < 32:

                self.driver.find_element(By.ID, "ctl00_MainContent_CldFecha_DdlDia").click()
                time.sleep(.5)

                d=expected_conditions.presence_of_element_located((By.XPATH,self.txtdia))
                
                WebDriverWait(self.driver, self.TEsperaMax).until(d)
                time.sleep(.5)
                #self.driver.find_element(By.XPATH,self.txtdia).click()
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFecha_DdlDia").send_keys(self.DIAF)
                self.driver.find_element(By.ID, "ctl00_MainContent_CldFecha_DdlDia").send_keys(Keys.ENTER)

                time.sleep(.5)

            else:
                pass

            self.driver.find_element(By.ID, "ctl00_MainContent_BtnBusqueda").click()
            time.sleep(1)
            
            try:
                meta=expected_conditions.presence_of_element_located((By.ID, "ListaFolios"))
                WebDriverWait(self.driver,15).until(meta)

                pass
            except:
                print("NO HAY CFDI")

                self.driver.quit()
                return

                    


            meta=expected_conditions.presence_of_element_located((By.ID, "ListaFolios"))
            WebDriverWait(self.driver,15).until(meta)

            time.sleep(.5)
            for i in range(2,1000):

                try:

#DESCARGA DE xml #id=BtnVerDetalle     id=BtnDescarga          id=BtnRI

                    me=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga"))
                    WebDriverWait(self.driver,2).until(me)
                        
                    time.sleep(.5)
                    VentanaPrincipal=self.driver.window_handles[0]
                    print(VentanaPrincipal)
                    time.sleep(1)

                    self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga").click()
                    time.sleep(5)

                    VentanaDescarga=self.driver.window_handles[1]
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaDescarga)
                    print(VentanaDescarga)

                    time.sleep(.5)

                    self.driver.close()
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaPrincipal)
                    time.sleep(.5)
                    print("CFDI XML Descargada  NUEMERO: " ,i-1)
                    time.sleep(1)

#decarga tambien el PDF
                    self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnRI").click()
                    time.sleep(2)

                    VentanaDescarga=self.driver.window_handles[1]
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaDescarga)
                    print(VentanaDescarga)

                    time.sleep(.5)

                    self.driver.close()
                    time.sleep(.5)

                    self.driver.switch_to.window(VentanaPrincipal)
                    time.sleep(.5)
                    print("CFDI PDF Descargada  NUEMERO: " ,i-1)
                    time.sleep(.5)






                    #self.driver.find_element(By.LINK_TEXT,"»").click()
                    

                      
                except:
                    try:
                        txtfac=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) > td:nth-child(2) > span"))
                        WebDriverWait(self.driver,.5).until(txtfac)

                        if txtfac:
                            print("existe CFDI:", i-1, "pero no encuentra el boton de descarga : CANCELADA ")

                    except:

                        print("No existe  CFDI:", i-1)
                        pass

                    

                if (i-1)%15==0:
                    try: 
                        self.driver.find_element(By.LINK_TEXT, "»").click()
                        time.sleep(.5)
                        self.driver.find_element(By.XPATH, '//body' ).send_keys(Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP,Keys.ARROW_UP)
                        #self.driver.find_element(By.CSS_SELECTOR, f"tr:nth-child({i}) > td:nth-child(2) > span").send_keys(Keys.ARROW_UP)
                        me=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, f"tr:nth-child({i}) #BtnDescarga"))
                        WebDriverWait(self.driver,1).until(me)
            
                        print("Siguiente página ") 

                        time.sleep(.5)

                    except:
                        print("No hay más CFDI")
                        break
                              


            print("Término descarga Individual de CFDI")
           
            
            
            #### DESCARGAR METADATA ####
            self.driver.find_element(By.ID, "ctl00_MainContent_BtnMetadata").click()
            time.sleep(1)

            #nombre del archivo zip a descargar  y comparar
            meta=expected_conditions.presence_of_element_located((By.XPATH , '//*[@id="dvAlert"]/div/strong'))
            WebDriverWait(self.driver, self.TEsperaMax).until(meta)
            
            tmeta=self.driver.find_element(By.CSS_SELECTOR, ".alert-success").text.strip()
            tmeta=tmeta[137:173]

            print(tmeta)
            time.sleep(.5)

            self.driver.get("https://portalcfdi.facturaelectronica.sat.gob.mx/Consulta.aspx")
            time.sleep(1)

            rd=expected_conditions.presence_of_element_located((By.LINK_TEXT, "Recuperar Descargas de CFDI"))
            WebDriverWait(self.driver, self.TEsperaMax).until(rd)
            time.sleep(.5)

            self.driver.find_element(By.LINK_TEXT, "Recuperar Descargas de CFDI").click()
            time.sleep(.5)
            
            try:
                zipc=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)"))
                WebDriverWait(self.driver, 10).until(zipc)
                time.sleep(.5)

            except:
                print("No hay descargas de CFDI ")
                self.driver.refresh()
            #try:
            #    zi=self.driver.find_element(By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)").text.strip()

            #except:
            #   self.driver.refresh()
            zi=0
            print(zi)
            contador=0

            while zi != tmeta:
                self.driver.refresh()
                time.sleep(5)
                try:
                    zipc=expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)"))
                    WebDriverWait(self.driver, 3).until(zipc)
                    time.sleep(.5)

                    zi=self.driver.find_element(By.CSS_SELECTOR, "tr:nth-child(2) > td:nth-child(2)").text.strip()

                except:
                    self.driver.refresh()
                contador=contador+1
                print(tmeta+"-- Mes:"+self.MESF+"  Intento: "+str(contador))
                print(zi)
                if zi == tmeta:
                    break




            if zi == tmeta:
                print(" Descarga Coincide a la consultada se procedea descargar ")
                pass
            else:
                self.driver.refresh()


            descargar=expected_conditions.presence_of_element_located((By.ID, "BtnDescarga"))
            WebDriverWait(self.driver, self.TEsperaMax).until(descargar)
            time.sleep(.5)

            self.driver.find_element(By.ID, "BtnDescarga").click()
            time.sleep(4)
                
        else:
            print(">-----<  TIPO DE PROCEDIMIENTO INCORRECTO ")
            pass
   
    def ConvExl(self):#,path):

        #self.path= path
        
        
       # self.df = pd.read_csv(self.path +"/*", sep='~')
       # self.df.to_excel('output.xlsx', 'EGRESOS')
        for i in os.listdir(self.path):
            if i.endswith(".zip"):
                try:
                    with zipfile.ZipFile(str(self.path)+"/"+i, 'r') as zip_ref:
                        zip_ref.extractall(self.path)
                        print("Se DESCOMPRIMIO:"+ i)
                        os.remove(str(self.path)+"/"+str(i))

                except:
                    print("No se pudo extraer ZIP de:"+i)
                    print(str(self.path)+"/"+i)
                    pass

            #print (i)

        for i in os.listdir(self.path):
            if i.endswith(".txt"):
                try:
                    self.df = pd.read_csv(self.path +"/"+i, sep='~')

                    self.df.to_excel(str(self.path+"/"+str(self.DIAF)+"-"+str(self.MESF)+"-"+str(self.ANOF)+"--"+str(self.TIPO)+"-"+str(self.NOMBRE)+".xlsx"),str(self.TIPO))
                    os.remove(self.path+"/"+i)
                    print("Generación exitosa de Excel.")
                    time.sleep(1)

                except:
                    print("No se generó Excel "+i)
                    pass

    def AddRegistro(self):
        excel=0
        #if not os.listdir(self.ClientePathRegAnual):
        for i in os.listdir(self.ClientePathRegAnual):
            print(i)
            if i.endswith(".xlsx"):
                excel=1
                print("ya existe Registro Anual de: "+self.NombreCompleto)
                pass
            else:
                pass
                
        if excel==0:

            shutil.copyfile(self.ClientePathRegEXMP, self.ClientePathRegAnual+"/"+self.NombreCompleto+str(self.ANIOACT)+".xlsx")


        print("Termino ADD REGISTRO")

    def TTDatExl(self):

        self.MONTOS=[]
        self.MONTOSVAL=[]
        self.SUBTOTALVAL=[]
        self.IVAVAL=[]
        self.TOTALVAL=[]

        self.SUMMONTOS=[]
        self.SUMSUBTOTAL=[]
        self.SUMIVA=[]
        self.SUMTOTAL=[]

        self.redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='lightUp')

        exl=str(self.path+"/"+str(self.DIAF)+"-"+str(self.MESF)+"-"+str(self.ANOF)+"--"+str(self.TIPO)+"-"+str(self.NOMBRE)+".xlsx")
        wb=openpyxl.load_workbook(exl) 
        self.wsheet = wb.active 
        self.wsheet.title =  str(self.TIPO)+"-"+str(self.NOMBRE)
        print("-----> Open excel")

        for fila in range(2,(self.wsheet.max_row)+1):
            for col in range(1,(self.wsheet.max_column)+1):
                self.Letra= get_column_letter(col)
                #print("ITERACION")

                if self.Letra == "H":
                    self.Fecha=self.wsheet[self.Letra+str(fila)].value
                    self.FechaF = str(self.ANIOF)+"-"+str(self.MESF)
                    ####match
                    m=(SM(None, self.Fecha, self.FechaF).ratio())
                    print(m)

                    if m < 0.52:
                        self.wsheet[self.Letra+str(fila)].fill=self.redFill
                        print(">-----< No es el mismo mes el Consultado y el Descargado")
                    else:
                        pass

                #MONTOS COLUMNA

                if self.Letra == "J":
                    self.MONTOS.append(self.wsheet[self.Letra + str(fila)].value)
                    valJ=self.wsheet[self.Letra + str(fila)].value
                    #print(valJ)

                    self.wsheet["O"+str(fila)]=valJ/1.16

                    valO=self.wsheet["O" + str(fila)].value
                    self.wsheet["P"+str(fila)]=valJ - valO

                    valP=self.wsheet["P" + str(fila)].value
                    self.wsheet["Q"+str(fila)]=valO + valP
                    
                    valQ=self.wsheet["Q" + str(fila)].value

                #efecto del comprobante si es INGRESO EGRESO NOMINA FATO PAGO
                elif self.Letra =="K":
                    valk=self.wsheet[self.Letra + str(fila)].value
                    if valk == "E":
                        valJ=self.wsheet["J" + str(fila)].value
                        valO=self.wsheet["O" + str(fila)].value
                        valP=self.wsheet["P" + str(fila)].value
                        #valQ=self.wsheet["Q" + str(fila)].value


                        self.wsheet["O"+str(fila)]=(valJ/1.16)*(-1)
                        self.wsheet["P"+str(fila)]=(valJ-valO)*(-1)
                        self.wsheet["Q"+str(fila)]=(valO + valP)*(-1)

                    elif valk == "N":
                        self.wsheet["O"+str(fila)]=0
                        self.wsheet["P"+str(fila)]=0
                        self.wsheet["Q"+str(fila)]=0
                #estado del comprobante       
                elif self.Letra =="L":
                    valL=self.wsheet[self.Letra + str(fila)].value
                    #si esta cancelada 
                    if valL == 0:
                        self.wsheet["O"+str(fila)]=0
                        self.wsheet["P"+str(fila)]=0
                        self.wsheet["Q"+str(fila)]=0

                else:
                    pass
                    

        for fila in range(2,(self.wsheet.max_row)+5):
            for col in range(1,(self.wsheet.max_column)+5):
                self.Letra= get_column_letter(col)

                #SUBTOTAL

        for fila in range(2,self.wsheet.max_row +1):# self.Letra =="O":
             self.SUBTOTALVAL.append(self.wsheet["O"+str(fila)].value)
             self.IVAVAL.append(self.wsheet["P"+str(fila)].value)
             self.TOTALVAL.append(self.wsheet["Q"+str(fila)].value)





        #montos VALORES NADAMAS  #ya no es necesario con ws.row_max y col_max          
        for i in self.MONTOS:
            if i != None:
                self.MONTOSVAL.append(i)
        

        try:        
            self.SUMMONTOS = round(sum(self.MONTOSVAL),1)
            self.SUMSUBTOTAL = round(sum(self.SUBTOTALVAL),1)
            self.SUMIVA = round(sum(self.IVAVAL),1)
            self.SUMTOTAL = round(sum(self.TOTALVAL),1)
        
        except:
            pass
                
        #colocar val en J de la suma de los montos
        self.wsheet["J"+str(self.wsheet.max_row+3)]="SÚMAS:  "
        self.wsheet["J"+str(self.wsheet.max_row+1)]=self.SUMMONTOS

        self.wsheet["O"+str(self.wsheet.max_row-1)]=self.SUMSUBTOTAL
        #self.wsheet["O"+str(self.wsheet.max_row)].fill=self.redFill
        self.wsheet["P"+str(self.wsheet.max_row-1)]=self.SUMIVA
        self.wsheet["Q"+str(self.wsheet.max_row-1)]=self.SUMTOTAL
        #encabezados OPQ        self.wsheet["J"+str(self.wsheet.max_row+3)]="SUMA NETO:  "+str(self.SUMMONTOS)
        self.wsheet["O"+str(self.wsheet.min_row)]=" SUBTOTAL:  "
        self.wsheet["P"+str(self.wsheet.min_row)]=" IVA:  "
        self.wsheet["Q"+str(self.wsheet.min_row)]=" TOTAL:  "

        #print(self.MONTOS)              
        #print(self.MONTOSVAL)              
        #print(self.SUMMONTOS)              
        print(self.SUMSUBTOTAL)              
        print(self.SUMIVA)              
        print(self.SUMTOTAL)              
    
        wb.save(exl)

        REG=str(self.ClientePathRegAnual+"/"+self.NombreCompleto+str(self.ANIOACT)+".xlsx")
        wb=openpyxl.load_workbook(REG) 
        self.wsheet = wb.active 
        self.SUMSUBTOTAL = sum(self.SUBTOTALVAL) 
        print("-----> Open Registro : "+str(self.TIPO).upper())
 #       print(self.TIPO)
        if self.TIPO =="INGRESO":
            if self.MESF=="01":
                self.wsheet["E"+str(6)]=int(self.SUMSUBTOTAL)

            elif self.MESF=="02":
                self.wsheet["E"+str(7)]=int(self.SUMSUBTOTAL)
            
            elif self.MESF=="03":
                self.wsheet["E"+str(8)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="04":
                self.wsheet["E"+str(9)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="05":
                self.wsheet["E"+str(10)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="06":
                self.wsheet["E"+str(11)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="07":
                self.wsheet["E"+str(12)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="08":
                self.wsheet["E"+str(13)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="09":
                self.wsheet["E"+str(14)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="10":
                self.wsheet["E"+str(15)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="11":
                self.wsheet["E"+str(16)]=int(self.SUMSUBTOTAL)
            elif self.MESF=="12":
                self.wsheet["E"+str(17)]=int(self.SUMSUBTOTAL)

            else:
                pass
        #else:
            #pass

        elif self.TIPO =="EGRESO":
            if self.MESF=="01":
                self.wsheet["F"+str(6)]=int(self.SUMSUBTOTAL)

            elif self.MESF=="02":
                self.wsheet["F"+str(7)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="03":
                self.wsheet["F"+str(8)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="04":
                self.wsheet["F"+str(9)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="05":
                self.wsheet["F"+str(10)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="06":
                self.wsheet["F"+str(11)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="07":
                self.wsheet["F"+str(12)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="08":
                self.wsheet["F"+str(13)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="09":
                self.wsheet["F"+str(14)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="10":
                self.wsheet["F"+str(15)]=int(self.SUMSUBTOTAL)

            
            elif self.MESF=="11":
                self.wsheet["F"+str(16)]=int(self.SUMSUBTOTAL)
            elif self.MESF=="12":
                self.wsheet["F"+str(17)]=int(self.SUMSUBTOTAL)

            else:
                pass
        else:
            print(">-----< Error: Tipo de comprobante. ")
            pass
            
        wb.save(REG)

    def LeerPDF(self):
        for i in os.listdir(self.path):
            if i.endswith(".pdf"):
                print(i)
                reader = PdfFileReader(open(f"{self.path}/{i}", 'rb'))
                print("si leyó PDF ")
                text = ""
                for page in reader.pages:
                    text += page.extract_text() 
                
                #text.find("Subtotal $")
                print(text)

    def LeerXML(self):
        self.XMLINFO=[]
        self.XMLSubtotal=[]
        self.XMLTotal=[]
        self.XMLTotalImpuestosTrasladados=[]
        self.XMLTotalImpuestosRetenidos=[]
        self.XMLImpuestosTrasladados=[]
        self.XMLMetodoPago=[]
        self.XMLUUID=[]

        self.XMLDescripcion=defaultdict(str)
        self.XMLValImpTrasladado=defaultdict(float)
        self.XMLValMonto=defaultdict(float)
        self.XMLValSubtotal=defaultdict(float)
        self.XMLValImpRet=defaultdict(float)

        self.LFacCanceladasTotalMontos=[]
        self.LFacCanceladasTotalImpTras=[]
        self.LFacCanceladasTotalSubtotal=[]
        self.LFacCanceladasTotalImpRet=[]


        self.LFacEgresoTotalMontos=[]
        self.LFacEgresoTotalImpTras=[]
        self.LFacEgresoTotalSubtotal=[]
        self.LFacEgresoTotalImpRet=[]

        self.LFacNominaTotalMontos=[]
        self.LFacNominaTotalImpTras=[]
        self.LFacNominaTotalSubtotal=[]
        self.LFacNominaTotalImpRet=[]





        self.FacCanceladas=0
        self.contador=0


        for i in os.listdir(self.path):
            if i.endswith(".xml"):
                print("Inic. reading XML")

                self.VdescStr="-" 
                self.VMonto=0
                self.VSubtotal=0
                self.VImpT=0
                self.VImpRet=0
                self.Vuuid=""
                self.Vdescripcion=""


                self.xml= self.path +"/"+i
                #mytree=minidom.parse(xml)
                self.tree=ET.parse(self.xml)
                self.root= self.tree.getroot()
                self.atribroot=self.root.iter()           

                for Atrib in self.atribroot:
                    #print(i.attrib)
                    ATRIBUTO=Atrib.attrib
                    self.XMLINFO.append(ATRIBUTO)
                    for k,v in ATRIBUTO.items():
                        #print("Inic. iter atributos XML")

                        if k =="MetodoPago":
                            self.XMLMetodoPago.append(v)
                            print(k,v)

                        elif k =="Total":
                            self.XMLTotal.append(float(v))
                            self.VMonto=float(v)
                            print(k,v)
                            
                        elif k =="SubTotal":
                            self.XMLSubtotal.append(float(v))
                            self.VSubtotal=float(v)
                            print(k,v)

#primero lee la descripcion por eso se añade en lUIDD
                        elif k =="Descripcion":
                            self.Vdescripcion=str(v)
                            self.VdescStr += (self.Vdescripcion+'----')
                            print(k,v)


                        elif k =="TotalImpuestosTrasladados":
                            self.XMLTotalImpuestosTrasladados.append(float(v))
                            self.VImpT=float(v)
                            print(k,v)

                        elif k =="TotalImpuestosRetenidos":
                            self.XMLTotalImpuestosRetenidos.append(float(v))
                            self.VImpRet=float(v)

                            print(k,v)

                        elif k =="UUID":

                            self.XMLUUID.append(str(v))
                            self.Vuuid=str(v)
                            self.XMLDescripcion[self.Vuuid]+=self.VdescStr
                            print(k,v)

                            try:
                                self.XMLValMonto[self.Vuuid]+=self.VMonto
                            
                            except:
                                print("")

                            try:
                                self.XMLValSubtotal[self.Vuuid]+=self.VSubtotal
                            except:
                                print("")

                            try:
                                self.XMLValImpRet[self.Vuuid]+=self.VImpRet
                            except:
                                print("")

                            try:
                                self.XMLValImpTrasladado[self.Vuuid]+=self.VImpT
                            except:
                                print("")

                        else:
                            pass

                    

                                


                self.contador+=1


        self.XMLTotalSuma=round(sum(self.XMLTotal),1)
        self.XMLSubtotalSuma=round(sum(self.XMLSubtotal),1)
        self.XMLTotalImpuestosTrasladadosSuma=round(sum(self.XMLTotalImpuestosTrasladados),1)
        self.XMLTotalImpuestosRetenidosSuma=round(sum(self.XMLTotalImpuestosRetenidos),1)

        exl=str(self.path+"/"+str(self.DIAF)+"-"+str(self.MESF)+"-"+str(self.ANOF)+"--"+str(self.TIPO)+"-"+str(self.NOMBRE)+".xlsx")
        wb=openpyxl.load_workbook(exl) 
        self.wsheet = wb.active 
        self.wsheet.title =  str(self.TIPO)+"-"+str(self.NOMBRE)
        #self.wsheet.column_dimensions['A:Z'].width = 30
        print("Se abrió el excel bien ")

        #leer Colocar concepto Leido de xml

        for fila in range(2,(self.wsheet.max_row)):
            for col in range(1,(self.wsheet.max_column)+10):
                self.Letra= get_column_letter(col)

                if self.Letra=="B":
                    ValB= self.wsheet[self.Letra + str(fila)].value 
                    #print("VALOR DE  COL B :" ,ValB )
                    if ValB!=None :

                        try:
                            self.wsheet["S" + str(self.wsheet.min_row)] = "Monto"
                            self.wsheet["S" + str(fila)] = self.XMLValMonto[ValB]
                        except:
                            print(">-----< ERR Monto")


                        try:
                            self.wsheet["T" +str( self.wsheet.min_row)] = "Imp Trasladados"
                            self.wsheet["T" + str(fila)] = self.XMLValImpTrasladado[ValB]

                        except:
                            print(">-----< ERR Imp. Trasladados")


                        try:
                            self.wsheet["U" +str( self.wsheet.min_row)] = "Imp Retenidos"
                            self.wsheet["U" + str(fila)] = self.XMLValImpRet[ValB]

                        except:
                            print(">-----< ERR Imp. Retenidos")


                        try:
                            self.wsheet["V" + str(self.wsheet.min_row)] = "Conceptos"
                            self.wsheet["V" + str(fila)] = self.XMLDescripcion[ValB]
                            

                        except:
                            print(">-----< Error en  concepto")



                elif self.Letra=="L":
                    ValL = self.wsheet[self.Letra +str(fila)].value
                    ValL=str(ValL)

                    if ValL=="0":
                        ValBFacCancelada=self.wsheet["B"+str(fila)].value


                        ValImpTras=self.XMLValImpTrasladado[ValBFacCancelada]
                        self.LFacCanceladasTotalImpTras.append(ValImpTras)

                        ValTotal=self.XMLValMonto[ValBFacCancelada]
                        self.LFacCanceladasTotalMontos.append(ValTotal)
                        
                        ValSubTotal=self.XMLValSubtotal[ValBFacCancelada]
                        self.LFacCanceladasTotalSubtotal.append(ValSubTotal)

                        ValImpRet=self.XMLValImpRet[ValBFacCancelada]
                        self.LFacCanceladasTotalImpRet.append(ValImpRet)

                        self.FacCanceladas+=1



                elif self.Letra=="K":
                    ValK = self.wsheet[self.Letra +str(fila)].value
                    ValK=str(ValK)

                    if ValK=="E":
                        ValBFacEgreso=self.wsheet["B"+str(fila)].value


                        ValImpTras=self.XMLValImpTrasladado[ValBFacEgreso]
                        self.LFacEgresoTotalImpTras.append(ValImpTras)

                        ValTotal=self.XMLValMonto[ValBFacEgreso]
                        self.LFacEgresoTotalMontos.append(ValTotal)
                        
                        ValSubTotal=self.XMLValSubtotal[ValBFacEgreso]
                        self.LFacEgresoTotalSubtotal.append(ValSubTotal)

                        ValImpRet=self.XMLValImpRet[ValBFacEgreso]
                        self.LFacEgresoTotalImpRet.append(ValImpRet)

                    elif ValK=="N":
                        ValBFacNomina=self.wsheet["B"+str(fila)].value


                        ValImpTras=self.XMLValImpTrasladado[ValBFacNomina]
                        self.LFacNominaTotalImpTras.append(ValImpTras)

                        ValTotal=self.XMLValMonto[ValBFacNomina]
                        self.LFacNominaTotalMontos.append(ValTotal)
                        
                        ValSubTotal=self.XMLValSubtotal[ValBFacNomina]
                        self.LFacNominaTotalSubtotal.append(ValSubTotal)

                        ValImpRet=self.XMLValImpRet[ValBFacNomina]
                        self.LFacNominaTotalImpRet.append(ValImpRet)

                else:
                    pass


        self.XMLSumaTotalMontosFacCanceladas=round(sum(self.LFacCanceladasTotalMontos),1)
        self.XMLSumaImpTrasFacCanceladas=round(sum(self.LFacCanceladasTotalImpTras),1)
        self.XMLSumaSubtotalesFacCanceladas=round(sum(self.LFacCanceladasTotalSubtotal),1)
        self.XMLSumaImpRetFacCanceladas=round(sum(self.LFacCanceladasTotalImpRet),1)

        self.XMLSumaTotalMontosFacEgreso=round(sum(self.LFacEgresoTotalMontos),1)
        self.XMLSumaImpTrasFacEgreso=round(sum(self.LFacEgresoTotalImpTras),1)
        self.XMLSumaSubtotalesFacEgreso=round(sum(self.LFacEgresoTotalSubtotal),1)
        self.XMLSumaImpRetFacEgreso=round(sum(self.LFacEgresoTotalImpRet),1)

        self.XMLSumaTotalMontosFacNomina=round(sum(self.LFacNominaTotalMontos),1)
        self.XMLSumaImpTrasFacNomina=round(sum(self.LFacNominaTotalImpTras),1)
        self.XMLSumaSubtotalesFacNomina=round(sum(self.LFacNominaTotalSubtotal),1)
        self.XMLSumaImpRetFacNomina=round(sum(self.LFacNominaTotalImpRet),1)


        self.GranTotalMonto= self.XMLTotalSuma + self.XMLSumaTotalMontosFacEgreso - self.XMLSumaTotalMontosFacCanceladas - self.XMLSumaTotalMontosFacNomina
        self.GranTotalSubtotal= self.XMLSubtotalSuma + self.XMLSumaSubtotalesFacEgreso - self.XMLSumaSubtotalesFacCanceladas - self.XMLSumaSubtotalesFacNomina
        self.GranTotalImpTras= self.XMLTotalImpuestosTrasladadosSuma + self.XMLSumaImpTrasFacEgreso - self.XMLSumaImpTrasFacCanceladas - self.XMLSumaImpTrasFacNomina
        self.GranTotalImpRet= self.XMLTotalImpuestosRetenidosSuma + self.XMLSumaImpRetFacEgreso - self.XMLSumaImpRetFacCanceladas - self.XMLSumaImpRetFacNomina

        print(self.GranTotalMonto)
        print(self.GranTotalSubtotal)
        print(self.GranTotalImpTras)
        print(self.GranTotalImpRet)



        for fila in range(1):
            for col in range(1,(self.wsheet.max_column)+10):
                self.Letra= get_column_letter(col)
                #print("ITERACION")

                if self.GranTotalMonto == self.SUMTOTAL:

                    if self.Letra == "O":
                        #deberia de ser solo IVA no total de impuestos trastladados ya que va iva +ieps ##esta  MAL
                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  CON IVA (16%): ." 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] = self.GranTotalImpTras /0.16

                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  SIN IVA : ." 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] =self.GranTotalMonto-(self.GranTotalImpTras /0.16)-self.GranTotalImpRet-self.GranTotalImpTras

                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  CON RETENCION (1.25%) : ." 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] = self.GranTotalImpRet /0.0125

                    elif self.Letra == "P":
                        self.wsheet[self.Letra + str(self.wsheet.max_row-7)] = "Imp.Trasladados. " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row-6)] = self.GranTotalImpTras

                        self.wsheet[self.Letra + str(self.wsheet.max_row-1)] = "Imp. Retenidos. " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row)] = self.GranTotalImpRet

                    elif self.Letra == "Q":
                        self.wsheet[self.Letra + str(self.wsheet.max_row-7)] = "TOTAL. " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row-6)] = self.GranTotalMonto

                    else:
                        pass


        
                else:
                    if self.Letra == "O":
                    
                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  Grabado (16%): " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] = "R:  "+str(self.GranTotalImpTras /0.16)
                    #if self.Letra == "T":
                        #deberia de ser solo IVA no total de impuestos trastladados ya que va iva +ieps ##esta  MAL
                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  SIN grabado : " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] = "R: "+str(self.GranTotalMonto-(self.GranTotalImpTras /0.16)-self.GranTotalImpRet)
                    
                    #if self.Letra == "U":
                        #deberia de ser solo IVA no total de impuestos trastladados ya que va iva +ieps ##esta  MAL
                        self.wsheet[self.Letra + str(self.wsheet.max_row+2)] = "SubTotal  CON RETENCION (1.25%): " 
                        self.wsheet[self.Letra + str(self.wsheet.max_row+1)] = "R: "+str(self.GranTotalImpRet /0.0125)

                    elif self.Letra == "P":
                        self.wsheet[self.Letra + str(self.wsheet.max_row-7)] = "Imp. Trasladados." 
                        self.wsheet[self.Letra + str(self.wsheet.max_row-6)] = self.GranTotalImpTras

                        self.wsheet[self.Letra + str(self.wsheet.max_row-1)] = "Imp. Retenidos." 
                        self.wsheet[self.Letra + str(self.wsheet.max_row)] = self.GranTotalImpRet

                    elif self.Letra == "Q":
                        self.wsheet[self.Letra + str(self.wsheet.max_row-7)] = "TOTAL" 
                        self.wsheet[self.Letra + str(self.wsheet.max_row-6)] = self.GranTotalMonto


                        


        for fila in range(1):
            for col in range(1,(self.wsheet.max_column)+1):
                self.Letra= get_column_letter(col)
            
                if self.Letra=="A":

                    self.wsheet["A" + str(self.wsheet.max_row+1)] = "# Metadata: "+str(self.wsheet.max_row-14)

                    self.wsheet["A" + str(self.wsheet.max_row+1)] = "# de XML Leidos : "+str(self.contador)
                    self.wsheet["A" + str(self.wsheet.max_row+1)] = "# CFDI. Canceladas : "+str(self.FacCanceladas)
                    self.wsheet["A" + str(self.wsheet.max_row+2)] = """
                    ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

                            Vargas Serrano & Asociados, Consultoría Fiscal y en Sistemas de Gestión de la Calidad ISO 9000. Agradece su preferencia.

                    ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
                    """
        wb.save(exl)

        """exl2=str(self.PATHSYS2+"/"+"InformeGral -"+str(self.ANOF)+"-"+str(self.MESF)+".xlsx")
        print("entro excel IGral")

        wb2=openpyxl.Workbook() 
        print("entro excel IGral")
        self.wsheet = wb2.active 
        print("entro excel IGral")
        self.wsheet.title =  "Informe General de Socios - Mes: "+str(self.MESF)
        print("entro excel IGral")

        for fila in range(1,(self.wsheet.max_row)+10):
            for col in range(1,(self.wsheet.max_column)+10):
                self.Letra= get_column_letter(col)
                #encabezados de Informe general 
                print("entro excel IGra interrrrrl")

                if self.Letra == "A":
                    self.wsheet[self.Letra+str(1)]= "NOMBRE DEL SOCIO."
                    print("entro excel IGra interrrrrl")

                    ValA=self.wsheet[self.Letra+str(fila)].value
                    print("entro excel IGra interrrrrl")

                    if ValA != self.NombreCompleto:
                        self.wsheet[self.Letra+str(3)]= self.NombreCompleto
                        print("entro excel IGra interrrrrl")




                elif self.Letra == "B":
                    self.wsheet[self.Letra+str(1)]= "Mes"

                
                elif self.Letra == "C":
                    self.wsheet[self.Letra+str(1)]= "CFDI INGRESO"
                    self.wsheet[self.Letra+str(1)]= "CFDI INGRESO"

                    #if self.TIPO == "INGRESO":


                elif self.Letra == "D":
                    self.wsheet[self.Letra+str(1)]= "TOTAL INGRESO"

                elif self.Letra == "E":
                    self.wsheet[self.Letra+str(1)]= "IMPUESTOS TRASLADADOS INGRESO"

                elif self.Letra == "F":
                    self.wsheet[self.Letra+str(1)]= "IMPUESTOS RETENIDOS INGRESO"

                else:
                    pass

        print("termino wb2")
        wb2.save(exl2)"""
   
    def InformeGeneral(self):
        for k,nom in DIC_NOMBRES.items:

            self.PathInformeGeneral=str(f"{self.PATHSYS2}/INFORME/{self.ANIOACT}/{self.MESF}")
            exl=pd.ExcelWriter(self.PathInformeGeneral, engine='xlsxwriter')#str(self.PathInformeGeneral+"/"+str(self.DIAACT)+"-"+str(self.MESACT)+"-"+str(self.ANIOACT)+".xlsx")
            wb=openpyxl.load_workbook(exl) 
            self.wsheet = wb.active 
            self.wsheet.title =  "Informe General de Socios - Mes"+str(self.MESF)

            for fila in range(2,(self.wsheet.max_row)+1):
                for col in range(1,(self.wsheet.max_column)+1):
                    self.Letra= get_column_letter(col)
                    #print("ITERACION")

                    if self.Letra == "H":
                        self.Fecha=self.wsheet[self.Letra+str(fila)].value
   
    def RUN(self):
            
        l1=list(DIC_NOMCLIENTSIEC.values())
        l2=list(DIC_NOMCLIENTEF.values())


        pathRUN=self.PATHSYS1

        for i in os.listdir(pathRUN):
            if i.endswith("RUN-IT-SAT.py"):
                f=open(pathRUN+ "/"+str(i),"r")
                textf=f.read()
                TextF=textf
                DicTexF= json.loads(TextF)
                print("Leyendo : ", i )
                print(DicTexF)

                for k,v in DicTexF.items():
                    if k=="Li":
                        Li=int(v)#int(input("Lugar en la lista inf: "))
                    
                    elif k== "Ls":
                        Ls=int(v)#15#int(input("Lugar en la lista sup: "))
                    
                    elif k=="L":   
                        L=int(v)

                    elif k== "O":    
                        O=str(v)#   Operación: CONSULTA,F3,F4: ")).upper()
                        
                    elif k== "T":
                        T=list(v)#["INGRESO"]#,"EGRESO"]#   TIPO: INGRESO, EGRESO: ")).upper()
                    
                    elif k== "D":
                        D=int(v)#int(input("Día: "))
                    
                    elif k== "MI":
                        MI=int(v)#int(input("Mes Inicio: "))
                    
                    elif k== "MF":
                        MF=int(v)+1#int(input("Mes Final: "))+1
                
                    elif k== "A":
                        A=int(v)#int(input("Año: "))

                    else:
                        pass

        for i in T:
            T=i
            for mes in range(MI,MF):
                Persona=""
                Tipo=str(T)
                anio=A
                #mes=8
                dia=D
                Operación=O #CONSULTA F4 F3 
                if mes >= 10:
                    mes=str(mes)
                else:
                    mes="0"+str(mes)

                if L==1:
                    for Persona in l1[Li:Ls]:
                        print(Persona,mes,Operación)
                        try:
                            T1=SAT(Persona,Tipo,anio,mes,dia,Operación)

                            #T1=SAT(str(Persona),Tipo,anio,mes,dia,Operación)#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                            if Operación=="CONSULTA":

                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 

                                T1.log()
                                #T1.logEF()
                                T1.Consultar()
                                #T1.LeerPDF()
                                T1.ConvExl()
                                T1.AddRegistro()


                                print("Termino convertir a excel")
                                T1.TTDatExl()
                                print("Termino TTDATOSl")
                                T1.LeerXML()
                                print("Termino LEER XML")

                                #input("quitar")
                                T1.quit()
                                print("Termino bien Consulta: " +Persona)

                            elif Operación=="F3":

                                print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                T1.fac3I()
                                T1.quit()
                                print("Termino bien F3: " +Persona)
                            elif Operación=="F4":

                                #T1=SAT(Persona,Tipo,anio,mes,dia,Operación)
                                print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                #T1.log()
                                T1.logEF()
                                #    def Fac40(self,FGlobal="NO",Periodicidad="Mensual",Mes="",ano="",TipoFactura="INGRESO",DescDetallada="Venta",ProdServ="",UnidadMedida="Unidad de servicio", Cantidad="1",VUnitario="0",Descuento="0",ObjetoImpuesto="si",NumIdentif="00",MultiplicadorConcepto=1,Moneda="Peso Mexicano",RfcCliente="",NommbreoRSCliente="",CpCliente="",RegFisCliente="",UsoFac="",FormaDePago="EFECTIVO", MetodoPago="PUE"):

                                T1.Fac40()
                                T1.quit()
                                print("Termino bien F4: " +Persona)

                            else:
                                print("TIPO DE CONSULTA ERRONEO")
                        except:
                            print(str(Persona) + ": NOTERMINO "+ str(mes)+ str(anio))
                            #T1.quit()
                            pass

                elif L==2:
                    for Persona in l2[Li:Ls]:
                        print(Persona,mes,Operación)
                        try:
                            T1=SAT(Persona,Tipo,anio,mes,dia,Operación)

                            #T1=SAT(str(Persona),Tipo,anio,mes,dia,Operación)#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                            if Operación=="CONSULTA":

                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 

                                #T1.log()
                                T1.logEF()
                                T1.Consultar()
                                T1.ConvExl()
                                T1.AddRegistro()


                                print("Termino convertir a excel")
                                T1.TTDatExl()
                                print("Termino TTDATOSl")
                                T1.LeerXML()
                                print("Termino LEERXMLS")

                                T1.quit()
                                print("Termino bien: " +Persona)

                            if Operación=="F3":

                                print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                T1.fac3I()
                                T1.quit()
                                print("Termino bien: " +Persona)
                            if Operación=="F4":

                                T1.Web()
                                T1.logEF()
                                T1.Fac40(
                                "NO",#FGlobal
                                "Mensual",#Periodicidad
                                "",#Mes
                                "",#ano
                                "Ingreso",#TipoFactura
                                "Consumo ",#DescDetallada
                                "cafeteria",#ProdServ
                                "Unidad de servicio",#UnidadMedida
                                "1",#Cantidad
                                "172.41",#VUnitario
                                "0",#Descuento
                                "si",#ObjetoImpuesto
                                "00",#NumIdentif
                                1,#MultiplicadorConcepto
                                "Peso Mexicano",#Moneda
                                "TRO010109A9A",#RfcCliente
                                "TARGET ROBOTICS",#NommbreoRSCliente
                                "78440",#CpCliente
                                "General de Ley Personas Morales",#RegFisCliente                     "General de Ley Personas Morales",
                                "Gastos en general",#UsoFac
                                "Efectivo", #FormaDePago
                                "Pago en una sola exhibición")#MetodoPago
                                T1.quit()
                                print("Termino bien: " +Persona)
                        except:
                            print(str(Persona) + ": NOTERMINO "+ str(mes)+ str(anio))
                            time.sleep(2)
                            T1.quit()
                            pass

                else:
                    print("LISTA de clientes INVÁLIDA")
      
    def APP(self):
        root=Tk()

        frame=Frame(root)#,bg="black")

        frame.pack(fill="both",expand="True")

        root.title("VARGAS SERRANO & ASOCIADOS -SAT- V:1.16")
        root.geometry('500x800+0+0')
        root.config(bg="white")

        Ins=Label(frame,text="Lista: 1 o 2").pack()
        E3=Entry(frame, width = 2,text="3")
        E3.insert(END,"2")
        E3.pack()
        
        Ins=Label(frame,text="# de Cliente ").pack()
        E1=Entry(frame, width = 2,text="1")
        E1.insert(END,"1")
        E1.pack()

        Ins=Label(frame,text="# de Cliente Final").pack()
        E2=Entry(frame, width = 2,text="2")
        E2.insert(END,"16")
        E2.pack()
        
        
        Ins=Label(frame,text="Operación : CONSULTA-F4").pack()
        E4=Entry(frame, width = 8,text="4")
        E4.insert(END,"CONSULTA")
        E4.pack()
        
        Ins=Label(frame,text='Tipo: INGRESO EGRESO').pack()
        E5=Entry(frame, width = 12,text="5")#,textvariable=["INGRESO","EGRESO"])
        E5.insert(END,"INGRESO EGRESO")
        E5.pack()
        
        #Ins=Label(frame,text="Dia").pack()
        #E6=Entry(frame, width = 2,text="6")
        #E6.insert(END,"0")
        #E6.pack()
        
        Ins=Label(frame,text="Mes Inicial").pack()
        E7=Entry(frame, width = 3,text="7")
        E7.insert(END,self.MESACT)
        E7.pack()
        
        Ins=Label(frame,text="Mes Final ").pack()
        E8=Entry(frame, width = 3,text="8")
        E8.insert(END,self.MESACT)
        E8.pack()
        
        Ins=Label(frame,text="Año").pack()
        E9=Entry(frame, width = 5,text="9")
        E9.insert(END,"2022")
        E9.pack()
        
        def RUN():
            Li=int(E1.get())-1
            Ls=int(E2.get())
            L=int(E3.get())
            O=E4.get().upper()
            #O=O.upper()
            T=E5.get().upper()
            T=list(T.split(" "))

            #D=int(E6.get())
            MI=int(E7.get())
            MF=int(E8.get())+1

            A=int(E9.get())

            l1=list(DIC_NOMCLIENTSIEC.values())
            l2=list(DIC_NOMCLIENTEF.values())
            print(Li,Ls,L,O,T,MI,MF,A)

            for i in T:
                T=i
                for mes in range(MI,MF):
                    Persona=""
                    Tipo=str(T)
                    anio=A
                    #mes=8
                    dia=0##################################################################### OJO SI NEECESITA DIA 
                    Operación=O #
                    if mes >= 10:
                        mes=str(mes)
                    elif mes<10:
                        mes="0"+str(mes)
                    else:
                        pass

                    if L==1:
                        for Persona in l1[Li:Ls]:
                            print(Persona,mes,Operación)
                            try:
                                T1=SAT(Persona,Tipo,anio,mes,dia,Operación)

                                #T1=SAT(str(Persona),Tipo,anio,mes,dia,Operación)#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                if Operación=="CONSULTA":

                                    T1.Web()#T1.ClientePathFE)#path para descargar archivos 

                                    T1.log()
                                    #T1.logEF()
                                    T1.Consultar()
                                    #T1.LeerPDF()
                                    T1.ConvExl()
                                    T1.AddRegistro()


                                    print("Termino convertir a excel")
                                    T1.TTDatExl()
                                    print("Termino TTDATOSl")
                                    T1.LeerXML()
                                    print("Termino LEER XML")

                                    #input("quitar")
                                    T1.quit()
                                    print("Termino bien Consulta: " +Persona)

                                elif Operación=="F3":

                                    print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                    T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                    T1.fac3I()
                                    T1.quit()
                                    print("Termino bien F3: " +Persona)
                                elif Operación=="F4":

                                    #T1=SAT(Persona,Tipo,anio,mes,dia,Operación)
                                    print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                    T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                    #T1.log()
                                    T1.logEF()
                                    #    def Fac40(self,FGlobal="NO",Periodicidad="Mensual",Mes="",ano="",TipoFactura="INGRESO",DescDetallada="Venta",ProdServ="",UnidadMedida="Unidad de servicio", Cantidad="1",VUnitario="0",Descuento="0",ObjetoImpuesto="si",NumIdentif="00",MultiplicadorConcepto=1,Moneda="Peso Mexicano",RfcCliente="",NommbreoRSCliente="",CpCliente="",RegFisCliente="",UsoFac="",FormaDePago="EFECTIVO", MetodoPago="PUE"):

                                    T1.Fac40()
                                    T1.quit()
                                    print("Termino bien F4: " +Persona)

                                else:
                                    print("TIPO DE CONSULTA ERRONEO")
                            except:
                                print(str(Persona) + ": NOTERMINO "+ str(mes)+ str(anio))
                                #T1.quit()
                                pass

                    elif L==2:
                        for Persona in l2[Li:Ls]:
                            print(Persona,mes,Operación)
                            try:
                                T1=SAT(Persona,Tipo,anio,mes,dia,Operación)

                                #T1=SAT(str(Persona),Tipo,anio,mes,dia,Operación)#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                if Operación=="CONSULTA":

                                    T1.Web()#T1.ClientePathFE)#path para descargar archivos 

                                    #T1.log()
                                    T1.logEF()
                                    T1.Consultar()
                                    T1.ConvExl()
                                    T1.AddRegistro()


                                    print("Termino convertir a excel")
                                    T1.TTDatExl()
                                    print("Termino TTDATOSl")
                                    T1.LeerXML()
                                    print("Termino LEERXMLS")

                                    T1.quit()
                                    print("Termino bien: " +Persona)

                                if Operación=="F3":

                                    print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                    T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                    T1.fac3I()
                                    T1.quit()
                                    print("Termino bien: " +Persona)
                                if Operación=="F4":

                                    T1.Web()
                                    T1.logEF()
                                    T1.Fac40()
                                    T1.quit()
                                    print("Termino bien: " +Persona)
                            except:
                                print(str(Persona) + ": NOTERMINO "+ str(mes)+ str(anio))
                                time.sleep(2)
                                T1.quit()
                                pass

                    else:
                        print("LISTA de clientes INVALIDA")
        
                
        def b2():
            time.sleep(30)
            print("Vargas Serrano & Asociados, agradece su preferencia.")
                     

        Boton=Button(frame,text="RUN-IT-SAT.",command=RUN,height = 2, width = 20,fg='green')#,state='disabled')
        Boton.pack()

        clear=Button(frame,text="V. S. & A.",command=b2,fg='white')#,state='disabled')
        clear.pack()

        text= Text(frame, wrap= WORD, font= ('Menlo 14'))
        text.insert(END,"""
1-'JUAN': 'JUAN'
2-'GERARDO_A':'GERARDO_A' ,
3-'EMILIO': 'EMILIO', 
4-'TIBERIO': 'TIBERIO', *
5-'JESUS': 'JESUS', *
6-'ANTONIO': 'ANTONIO', 
7-'MARIELA':'MARIELA',
8-'ALEJANDRO':'ALEJANDRO',
""")
        text.place(x=0, y= 530, width= 250, height= 250)


        text2= Text(frame, wrap= WORD, font= ('Menlo 14 '))
        text2.insert(END,"""
1-'LUCERO': 'LUCERO', 
2-'ALFREDO': 'ALFREDO',
3-'GERARDO': 'GERARDO',
4-'GUILLERMO': 'GUILLERMO',
5-'SANTIAGO': 'SANTIAGO',
6-'OSCAR':'OSCAR',
7-'PATRICIA':'PATRICIA',
8-'MANUEL':'MANUEL',
9-'RODRIGO':'RODRIGO',
10-'EMIGDIO' : 'EMIGDIO',
11-'DANIELA':'DANIELA',)
12-'SARAI': 'SARAI',
13-'ITALO': 'ITALO',""")
        text2.place(x=250, y= 530, width= 250, height= 250)        

        root.mainloop()


def Whats():
    

    FinMes = "Estimado RIF, RESICO y persona Física, con el gusto de saludarle y deseando se encuentre bien, el motivo de este mensaje es infórmale estamos próximos a fin de mes, por lo que nuestro equipo de trabajo estará realizando el recorrido correspondiente el día de mañana Jueves 29 de septiembre por la tarde, para recolectar los tickets de gastos pendientes por facturar así como el libro de ingresos. \n \n \nDeseándole excelente noche, quedamos a sus órdenes.  \nDespacho fiscal Vargas Serrano & Asociados "

    Recibido=' Buen Día, confirmo de recibido, saludos!'
    InformaciónImportante=' Se informa a nuestros clientes................ .\nDeseandole un excelente día, reciba un cordial saludo.'

    Contador=0



    for k,v in DIC_CEL.items():
        N=datetime.datetime.now()
        HNow=N.hour
        MNow=int(N.minute )
        #print(HN,MN)
        Contador=Contador+1
        print(Contador)
        print(HNow,MNow)
        pywhatkit.sendwhatmsg(v, FinMes, HNow, MNow+1)
        #pywhatkit.sendwhatmsg_instantly(v, InformaciónImportante, 1)

def APP():
    root=Tk()

    frame=Frame(root)#,bg="black")

    frame.pack(fill="both",expand="True")

    root.title(" - VARGAS SERRANO & ASOCIADOS - SAT - V:1.29")
    root.geometry('500x810+0+0')
    root.config(bg="white")
    Ins=Label(frame,text="Lista: 1 o 2\n").pack()
    E3=Entry(frame, width = 2,text="3")
    E3.insert(END,"2")
    E3.pack()
    
    Ins=Label(frame,text="# de Cliente ").pack()
    E1=Entry(frame, width = 2,text="1")
    E1.insert(END,"1")
    E1.pack()

    Ins=Label(frame,text="# de Cliente Final").pack()
    E2=Entry(frame, width = 2,text="2")
    E2.insert(END,"16")
    E2.pack()
    
    
    Ins=Label(frame,text="Operación : CONSULTA / F4").pack()
    E4=Entry(frame, width = 8,text="4")
    E4.insert(END,"CONSULTA")
    E4.pack()
    
    Ins=Label(frame,text='Tipo: INGRESO EGRESO').pack()
    E5=Entry(frame, width = 12,text="5")#,textvariable=["INGRESO","EGRESO"])
    E5.insert(END,"INGRESO EGRESO")
    E5.pack()
    
    Ins=Label(frame,text="Día").pack()
    E6=Entry(frame, width = 2,text="6")
    E6.insert(END,"0")#datetime.date.today().day)
    E6.pack()
    
    Ins=Label(frame,text="Mes").pack()
    E7=Entry(frame, width = 3,text="7")
    E7.insert(END,datetime.date.today().month)
    E7.pack()
    
    Ins=Label(frame,text="Mes Final (solo Consulta)").pack()
    E8=Entry(frame, width = 3,text="8")
    E8.insert(END,datetime.date.today().month)
    E8.pack()
    
    Ins=Label(frame,text="Año").pack()
    E9=Entry(frame, width = 5,text="9")
    E9.insert(END,"2022")
    E9.pack()
   

    def RUN():
        Li=int(E1.get())-1
        Ls=int(E2.get())
        L=int(E3.get())
        O=E4.get().upper()
        T=E5.get().upper()
        T=list(T.split(" "))
        try:
            D=int(E6.get())
        except:
            D=E6.get()

        MI=int(E7.get())
        MF=int(E8.get())+1

        A=int(E9.get())

        l1=list(DIC_NOMCLIENTSIEC.values())
        l2=list(DIC_NOMCLIENTEF.values())
        #print(str("Linferior:"+Li,"Lsuperior:"+Ls,"LLista:"+L,"Operación:"+O,"Tipo"+T,"Mes Inicial:"+MI,"Mesfinal:"+MF,"Año:"+A))

        for i in T:
            T=i
            for mes in range(MI,MF):
                Persona=""
                Tipo=str(T)
                anio=A
                #mes=8
                dia=D
                Operación=O #
                if mes >= 10:
                    mes=str(mes)
                else:
                    mes="0"+str(mes)

                if L==1:
                    for Persona in l1[Li:Ls]:
                        try:
                            T1=SAT(Persona,Tipo,anio,mes,dia,Operación)
                            if Operación=="CONSULTA":

                                T1.Web()
                                T1.log()
                                T1.Consultar()
                                T1.ConvExl()
                                T1.AddRegistro()
                                print("Termino convertir a excel")
                                T1.TTDatExl()
                                print("Termino TTDATOSl")
                                T1.LeerXML()
                                print("Termino LEER XML")
                                T1.quit()
                                print("Termino bien Consulta: " +Persona)
                                time.sleep(2)

                            elif Operación=="F3":

                                print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                T1.fac3I()
                                T1.quit()
                                print("Termino bien F3: " +Persona)
                            elif Operación=="F4":
                                

                                print("NO PUEDES PORQUE NO TIENES EFIRMA  ERES Lista 1")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, 
                                time.sleep(5)

                            else:
                                print("TIPO DE CONSULTA ERRONEO")
                        except:
                            print(">-----<")
                            print(">-------------<")
                            print(str(Persona) + ": NO TERMINÓ Satisfactoriamente  "+ str(mes)+ str(anio))
                            print(">-------------<")
                            print(">-----<")


                            T1.quit()
                            pass

                elif L==2:
                    for Persona in l2[Li:Ls]:
                        try:
                            T1=SAT(Persona,Tipo,anio,mes,dia,Operación)

                            if Operación=="CONSULTA":

                                T1.Web()

                                T1.logEF()
                                T1.Consultar()
                                T1.ConvExl()
                                T1.AddRegistro()
                                print("Termino convertir a excel")
                                T1.TTDatExl()
                                print("Termino TTDATOSl")
                                T1.LeerXML()
                                print("Termino LEERXMLS")
                                T1.quit()
                                print("Termino bien: " +Persona)
                                print("--------------------------------------------------")

                            if Operación=="F3":

                                print("si entra aqui ")#Ingreso o egreso;ANIO,MESDIA pra la facturacion, CONSULTA F3 F4
                                T1.Web()#T1.ClientePathFE)#path para descargar archivos 
                                T1.fac3I()
                                T1.quit()
                                print("Termino bien: " +Persona)
                            if Operación=="F4":

                                T1.Web()
                                T1.logEF()
                                time.sleep(1)
                                T1.Fac40()
                                print("Termino bien: " +Persona)
                                time.sleep(1)
                                T1.quit()
                                print("Termino bien: " +Persona)

                        except:
                            print(">-----<")
                            print(">-------------<")
                            print(str(Persona) + ": NO TERMINÓ Satisfactoriamente  "+ str(mes)+ str(anio)+str(Tipo)+str(Operación))
                            print(">-------------<")
                            print(">-----<")

                            time.sleep(2)
                            T1.quit()
                            pass

                else:
                    print("LISTA de clientes INVALIDA   1 = Clientes con SIEC   ;  2 = Clientes con eFirma")
    
            



    def b2():
        #Whats()
        print("Vargas Serrano & Asociados, agradece  su preferencia.")
        time.sleep(1)
        quit()

    Boton=Button(frame,text="--  RUN.  --",command=RUN,height = 2, width = 20,fg='green')#,state='disabled')
    Boton.pack()

    clear=Button(frame,text="V.S. & A.",command=b2,fg='white')#,state='disabled')
    clear.pack()

    termf = Frame(root, height=380, width=500)
#abrir  una terminal 
    """termf.pack(fill=BOTH, expand=YES)
    wid = termf.winfo_id()
    os.system('xterm -into %d -geometry 40x20 -sb &' % wid)"""


    text= Text(frame, wrap= WORD, font= ('Menlo 14'))
    text.insert(END,"""
    Lista 1: 

'1 -  JUAN': 'JUAN',
'2 -  GERARDO_A':'GERARDO_A' ,
'3 -  EMILIO': 'EMILIO', 
'4 -  TIBERIO': 'TIBERIO', 
'5 -  JESUS': 'JESUS', *
'6 -  ANTONIO': 'ANTONIO', 
'7 -  MARIELA':'MARIELA',
'8 -  ALEJANDRO':'ALEJANDRO',
""")
    text.place(x=0, y= 540, width= 250, height= 250)



    text2= Text(frame, wrap= WORD, font= ('Menlo 14 '))
    text2.insert(END,"""
    Lista 2: 

'1 -  LUCERO': 'LUCERO', 
'2 -  ALFREDO': 'ALFREDO',
'3 -  GERARDO': 'GERARDO',
'4 -  GUILLERMO': 'GUILLERMO',
'5 -  SANTIAGO': 'SANTIAGO',
'6 -  OSCAR':'OSCAR',
'7 -  PATRICIA':'PATRICIA',
'8 -  MANUEL':'MANUEL',
'9 -  RODRIGO':'RODRIGO',
'10 -  EMIGDIO' : 'EMIGDIO',
'11 -  DANIELA':'DANIELA',
'12 -  SARAI': 'SARAI',
'13 -  ITALO': 'ITALO',""")

    text2.place(x=250, y= 540, width= 250, height= 250)  

    root.mainloop()


if __name__=='__main__':
    APP()
        

